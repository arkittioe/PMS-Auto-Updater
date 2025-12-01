import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re
import json
import os
from datetime import datetime
import win32com.client
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any

import json


# ================================================================================
# 🔧 بارگذاری تنظیمات از config.json
# ================================================================================

class ConfigLoader:
    """کلاس بارگذاری تنظیمات از فایل JSON یا ENV"""

    @staticmethod
    def load_json(config_file: str = "config.json") -> Dict[str, Any]:
        """
        بارگذاری فایل JSON یا ENV (با سازگاری)

        Args:
            config_file: مسیر فایل config

        Returns:
            دیکشنری تنظیمات
        """
        # اگه فایل مستقیم موجوده، بارگذاری کن
        if os.path.exists(config_file):
            # چک کردن نوع فایل
            if config_file.endswith('.json'):
                return ConfigLoader._load_json_file(config_file)
            else:
                # احتمالا .env هست
                print(f"⚠️  استفاده از {config_file}")
                print(f"💡 توصیه: فایل را به config.json تبدیل کنید")
                return ConfigLoader._load_env_file(config_file)

        # فایل مستقیم نبود، بررسی فایل‌های جایگزین
        if config_file.endswith('.json'):
            # اگه JSON خواسته ولی نبود، دنبال .env بگرد
            env_file = config_file.replace('.json', '.env')
            if os.path.exists(env_file):
                print(f"⚠️  فایل {config_file} یافت نشد، از {env_file} استفاده می‌شود")
                print(f"💡 توصیه: فایل config.json بسازید (الگو در مستندات)")
                return ConfigLoader._load_env_file(env_file)
        elif config_file.endswith('.env'):
            # اگه .env خواسته ولی نبود، دنبال JSON بگرد
            json_file = config_file.replace('.env', '.json')
            if os.path.exists(json_file):
                print(f"✅ استفاده از {json_file} به جای {config_file}")
                return ConfigLoader._load_json_file(json_file)

        # هیچکدوم پیدا نشد
        raise FileNotFoundError(
            f"❌ فایل تنظیمات یافت نشد: {config_file}\n"
            f"💡 لطفا یکی از فایل‌های زیر را ایجاد کنید:\n"
            f"   - config.json (توصیه می‌شود)\n"
            f"   - config.env (سازگاری با نسخه قدیم)"
        )

    @staticmethod
    def _load_json_file(file_path: str) -> Dict[str, Any]:
        """بارگذاری فایل JSON"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            print(f"✅ تنظیمات از {file_path} بارگذاری شد")
            return config
        except json.JSONDecodeError as e:
            raise ValueError(f"❌ خطا در خواندن JSON از {file_path}: {e}")

    @staticmethod
    def _load_env_file(file_path: str) -> Dict[str, Any]:
        """بارگذاری فایل ENV"""
        config = {}

        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                if '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip()
                    value = value.strip()

                    # تبدیل به نوع صحیح
                    if value.lower() == 'true':
                        value = True
                    elif value.lower() == 'false':
                        value = False
                    elif value.isdigit():
                        value = int(value)
                    elif ',' in value:
                        value = [int(x.strip()) for x in value.split(',')]

                    config[key] = value

        print(f"✅ تنظیمات از {file_path} بارگذاری شد (ENV format)")
        return config

    @staticmethod
    def _convert_env_to_dict(env_file: str) -> Dict[str, Any]:
        """تبدیل .env به ساختار dictionary شبیه JSON"""
        config = {}

        with open(env_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                if '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip()
                    value = value.strip()

                    # تبدیل به نوع صحیح
                    if value.lower() == 'true':
                        value = True
                    elif value.lower() == 'false':
                        value = False
                    elif value.isdigit():
                        value = int(value)
                    elif ',' in value:
                        value = [int(x.strip()) for x in value.split(',')]

                    config[key] = value

        return config


# ================================================================================
# 🎛️ تنظیمات مرکزی (با پشتیبانی از JSON)
# ================================================================================

class PMSConfig:
    """
    کلاس تنظیمات مرکزی برنامه
    تمام مسیرها، نام‌ها و mapping ستون‌ها در اینجا تعریف می‌شوند
    """

    def __init__(self, config_file: str = "config.json"):
        """
        مقداردهی اولیه با بارگذاری از فایل config

        Args:
            config_file: مسیر فایل تنظیمات (JSON یا ENV)
        """
        config_data = ConfigLoader.load_json(config_file)

        # بررسی نوع config (JSON ساختاریافته یا ENV ساده)
        if 'files' in config_data:
            # JSON ساختاریافته
            self._load_from_json(config_data)
        else:
            # ENV ساده (سازگاری با نسخه قدیم)
            self._load_from_env_dict(config_data)

    def _load_from_json(self, config: Dict):
        """بارگذاری از JSON ساختاریافته"""
        # فایل‌ها و شیت‌ها
        self.PMS_FILE = config['files']['pms']['file']
        self.PMS_SHEET = config['files']['pms']['sheet']
        self.PNT_FILE = config['files']['pnt']['file']
        self.PNT_SHEET = config['files']['pnt']['sheet']
        self.CACHE_FILE = config['files']['cache']['file']

        # محدوده محورها
        self.AXIS_RANGE_START = config['axis']['range_start']
        self.AXIS_RANGE_END = config['axis']['range_end']

        # Cache
        self.USE_CACHE = config['files']['cache']['enabled']

        # ستون‌های PNT
        self.PNT = self._PNTConfig(config['columns']['pnt'])

        # ستون‌های PMS
        self.PMS = self._PMSConfig(config['columns']['pms'])

        # سلسله مراتب
        self.Hierarchy = self._HierarchyConfig(config['hierarchy'], self)

    def _load_from_env_dict(self, env_config: Dict):
        """بارگذاری از ENV dictionary (سازگاری)"""
        # فایل‌ها و شیت‌ها
        self.PMS_FILE = env_config.get('PMS_FILE', "PMS-paint REV-03H.xlsx")
        self.PMS_SHEET = env_config.get('PMS_SHEET', "1404.01.22")
        self.PNT_FILE = env_config.get('PNT_FILE', "PNT-G.xlsx")
        self.PNT_SHEET = env_config.get('PNT_SHEET', " المان PNT-G-130")
        self.CACHE_FILE = env_config.get('CACHE_FILE', "pms_cache.json")

        # محدوده محورها
        self.AXIS_RANGE_START = env_config.get('AXIS_START', 19)
        self.AXIS_RANGE_END = env_config.get('AXIS_END', 46)

        # Cache
        self.USE_CACHE = env_config.get('USE_CACHE', True)

        # ستون‌های PNT
        self.PNT = self._PNTConfig(env_config)

        # ستون‌های PMS
        self.PMS = self._PMSConfig(env_config)

        # سلسله مراتب
        self.Hierarchy = self._HierarchyConfig(env_config, self)

    class _PNTConfig:
        def __init__(self, config):
            if 'item' in config:  # JSON
                self.ITEM_COL = config['item']
                self.QUANTITY_COL = config['quantity']
                self.M_VALUE_COL = config['m_value']
                self.G2_CELL = config['g2_cell']
                self.ROW_AUTO = config.get('row_auto', True)
                self.AXIS_SEARCH_COLS = config['axis_search']
            else:  # ENV
                self.ITEM_COL = config.get('PNT_ITEM_COL', 3)
                self.QUANTITY_COL = config.get('PNT_QUANTITY_COL', 9)
                self.M_VALUE_COL = config.get('PNT_M_VALUE_COL', 13)
                self.G2_CELL = config.get('PNT_G2_CELL', "G2")
                self.ROW_AUTO = config.get('PNT_ROW_AUTO', True)
                self.AXIS_SEARCH_COLS = config.get('PNT_AXIS_SEARCH_COLS', [3, 4, 5])

            self.ROW_START = 7
            self.ROW_END = 31

    class _PMSConfig:
        def __init__(self, config):
            if 'text' in config:  # JSON
                self.TEXT_COL = config['text']
                self.DATE_COL = config['date']
                self.G_COL = config['g']
                self.N_COL = config['n']
            else:  # ENV
                self.TEXT_COL = config.get('PMS_TEXT_COL', 1)
                self.DATE_COL = config.get('PMS_DATE_COL', 5)
                self.G_COL = config.get('PMS_G_COL', 7)
                self.N_COL = config.get('PMS_N_COL', 14)

    class _HierarchyConfig:
        def __init__(self, config, parent_config):
            if 'level_1_pattern' in config:  # JSON
                self.LEVEL_1_PATTERN = config['level_1_pattern']
                self.LEVEL_3_TEXT = config['level_3_text']
                self.LEVEL_4_TEXT = config['level_4_text']
                self.TARGET_LEVEL = config['target_level']
            else:  # ENV
                self.LEVEL_1_PATTERN = config.get('HIERARCHY_LEVEL_1_PATTERN', "محور")
                self.LEVEL_3_TEXT = config.get('HIERARCHY_LEVEL_3_TEXT', "GLASS FLAKE")
                self.LEVEL_4_TEXT = config.get('HIERARCHY_LEVEL_4_TEXT',
                                               "بلاست و اماده سازی سطح  و اعمال رنگ  لایه دوم")
                self.TARGET_LEVEL = config.get('HIERARCHY_TARGET_LEVEL', 5)

            self.parent = parent_config

        def get_search_path(self, mohor_num: int) -> List[Tuple[int, str]]:
            return [
                (1, f"{self.LEVEL_1_PATTERN} {mohor_num}"),
                (3, self.LEVEL_3_TEXT),
                (4, self.LEVEL_4_TEXT)
            ]


# ================================================================================
# 🔧 ماژول نرمال‌سازی متن
# ================================================================================

class TextNormalizer:
    """
    کلاس نرمال‌سازی متن
    شامل متدهای مختلف برای نرمال‌سازی متن فارسی و انگلیسی
    """

    @staticmethod
    def clean_g2_value(text: Optional[str]) -> str:
        """
        پاکسازی مقدار G2 برای استفاده در PMS
        - حذف عبارت "شماره صورتمجلس"
        - حذف فاصله‌های اضافی

        Args:
            text: متن ورودی از سلول G2

        Returns:
            متن پاکسازی شده
        """
        if text is None:
            return ""

        text = str(text).strip()

        # حذف عبارت "شماره صورتمجلس" (با حروف مختلف ی/ی و ک/ک)
        text = re.sub(r'شماره\s*صورت\s*مجلس', '', text, flags=re.IGNORECASE)
        text = re.sub(r'شماره\s*صورتمجلس', '', text, flags=re.IGNORECASE)

        # حذف فاصله‌های اضافی
        text = ' '.join(text.split())

        return text.strip()

    @staticmethod
    def normalize_standard(text: Optional[str]) -> str:
        """
        نرمال‌سازی استاندارد برای متن فارسی

        Args:
            text: متن ورودی

        Returns:
            متن نرمال‌شده (lowercase, بدون فاصله)
        """
        if text is None:
            return ""

        text = str(text).strip()
        text = text.replace('ی', 'ی').replace('ک', 'ک')
        text = re.sub(r'\s+', '', text)
        text = text.lower()

        return text

    @staticmethod
    def normalize_pnt_axis(text: Optional[str]) -> str:
        """
        نرمال‌سازی پیشرفته برای استخراج محور از PNT-G
        - تبدیل \n به فاصله
        - تبدیل به uppercase
        - حذف فاصله‌ها و خط‌تیره

        Args:
            text: متن ورودی

        Returns:
            متن نرمال‌شده (UPPERCASE, بدون فاصله و خط‌تیره)
        """
        if text is None:
            return ""

        text = str(text)
        # تبدیل خطوط به فاصله
        text = text.replace('\n', ' ').replace('\r', ' ')
        # تبدیل به حروف بزرگ
        text = text.upper()
        # حذف فاصله‌ها و خط‌تیره
        text = text.replace(' ', '').replace('-', '')

        return text

    @staticmethod
    def multiline_to_single(text: Optional[str]) -> str:
        """
        تبدیل متن چندخطی به تک‌خطی
        - تبدیل \n و \r به فاصله
        - حذف فاصه‌های اضافی

        Args:
            text: متن چندخطی

        Returns:
            متن تک‌خطی
        """
        if text is None:
            return ""

        text = str(text)
        # تبدیل خطوط به فاصله
        text = text.replace('\n', ' ').replace('\r', ' ')
        # حذف فاصله‌های متوالی
        text = ' '.join(text.split())

        return text.strip()

# ================================================================================
# 🔧 ماژول استخراج محور
# ================================================================================

class AxisExtractor:
    """
    کلاس استخراج شماره محور از سطرهای PNT-G
    با دو اولویت جستجو: AXIS[19-45] و سپس S[19-45]
    """

    def __init__(self, config: PMSConfig):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.normalizer = TextNormalizer()

    def extract_from_row(self, ws: Worksheet, row_idx: int) -> Optional[int]:
        """
        استخراج شماره محور از یک سطر PNT-G

        جستجو در ستون‌های مشخص شده با دو اولویت:
        1. AXIS[19-45]
        2. S[19-45]

        Args:
            ws: worksheet PNT-G
            row_idx: شماره سطر

        Returns:
            شماره محور (19-45) یا None
        """
        # اولویت 1: جستجوی AXIS[19-45]
        mohor = self._search_pattern(ws, row_idx, "AXIS")
        if mohor:
            return mohor

        # اولویت 2: جستجوی S[19-45]
        mohor = self._search_pattern(ws, row_idx, "S")
        return mohor

    def _search_pattern(self, ws: Worksheet, row_idx: int, pattern_prefix: str) -> Optional[int]:
        """
        جستجوی الگو در ستون‌های مشخص

        Args:
            ws: worksheet
            row_idx: شماره سطر
            pattern_prefix: پیشوند الگو (AXIS یا S)

        Returns:
            شماره محور یا None
        """
        for col in self.config.PNT.AXIS_SEARCH_COLS:
            cell_value = ws.cell(row_idx, col).value
            if not cell_value:
                continue

            normalized = self.normalizer.normalize_pnt_axis(cell_value)

            for mohor_num in range(self.config.AXIS_RANGE_START,
                                   self.config.AXIS_RANGE_END):
                pattern = f"{pattern_prefix}{mohor_num}"
                if pattern in normalized:
                    return mohor_num

        return None

# ================================================================================
# 💾 ماژول مدیریت Cache
# ================================================================================

class PMSCacheManager:
    """کلاس مدیریت cache ساختار PMS"""

    def __init__(self, cache_file: str, log_callback=None):
        self.cache_file = cache_file
        self.log_callback = log_callback or print

    def load_cache(self, file_path: str, sheet_name: str) -> Optional[Dict]:
        """بارگذاری cache در صورت معتبر بودن"""
        if not os.path.exists(self.cache_file):
            return None

        try:
            with open(self.cache_file, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)

            current_hash = self.get_file_hash(file_path)

            if (cache_data.get('file_hash') == current_hash and
                    cache_data.get('sheet_name') == sheet_name):
                return cache_data['item_locations']
        except Exception as e:
            self.log_callback(f"⚠️  خطا در خواندن Cache: {e}", "warning")

        return None

    def save_cache(self, file_path: str, sheet_name: str, item_locations: Dict):
        """ذخیره cache"""
        try:
            cache_dir = os.path.dirname(self.cache_file)
            if cache_dir and not os.path.exists(cache_dir):
                os.makedirs(cache_dir)

            cache_data = {
                'file_hash': self.get_file_hash(file_path),
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'item_locations': item_locations
            }

            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)

            self.log_callback(f"💾 Cache ذخیره شد: {self.cache_file}", "info")
        except Exception as e:
            self.log_callback(f"⚠️  خطا در ذخیره Cache: {e}", "warning")

    @staticmethod
    def get_file_hash(file_path: str) -> str:
        """
        محاسبه hash ساده بر اساس modified time و size

        Args:
            file_path: مسیر فایل

        Returns:
            hash string
        """
        stat = os.stat(file_path)
        return f"{stat.st_mtime}_{stat.st_size}"

# ================================================================================
# 📊 ماژول خواندن ساختار PMS
# ================================================================================

class PMSStructureReader:
    """کلاس خواندن و استخراج ساختار سلسله‌مراتبی PMS"""

    def __init__(self, config: PMSConfig, log_callback=None):
        self.config = config
        self.log_callback = log_callback or print
        self.normalizer = TextNormalizer()
        self.hierarchy_searcher = PMSHierarchySearcher(config)



    def extract_all_items(self, file_path: str, sheet_name: str) -> Dict[str, List[Dict]]:
        """
        استخراج تمام آیتم‌های Level 5 از همه محورها

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت

        Returns:
            دیکشنری: {normalized_item: [{'mohor': ..., 'row': ..., 'level': ...}]}
        """
        self.log_callback("🔄 در حال استخراج ساختار PMS...", "info")

        # جستجو در تمام محورها
        mohor_results = self._search_all_mohors(file_path, sheet_name)

        # تبدیل به فرمت دیکشنری برای جستجوی سریع
        item_locations = {}
        for mohor_name, items in mohor_results.items():
            for item in items:
                normalized_text = self.normalizer.normalize_standard(item['text'])
                if normalized_text not in item_locations:
                    item_locations[normalized_text] = []
                item_locations[normalized_text].append({
                    'mohor': mohor_name,
                    'row': item['row'],
                    'level': item['level'],
                    'original_text': item['text']
                })

        return item_locations

    def _search_all_mohors(self, file_path: str, sheet_name: str) -> Dict[str, List[Dict]]:
        """
        جستجوی سلسله‌مراتبی در تمام محورها

        Args:
            file_path: مسیر فایل
            sheet_name: نام شیت

        Returns:
            دیکشنری: {mohor_name: [items]}
        """
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]

        all_results = {}

        for mohor_num in range(self.config.AXIS_RANGE_START,
                               self.config.AXIS_RANGE_END):
            mohor_name = f"محور {mohor_num}"
            search_path = self.config.Hierarchy.get_search_path(mohor_num)

            results = self.hierarchy_searcher.find_items(
                ws,
                search_path,
                self.config.Hierarchy.TARGET_LEVEL
            )

            if results:
                all_results[mohor_name] = results

        wb.close()
        return all_results

# ================================================================================
# 📊 ماژول جستجوی سلسله‌مراتبی
# ================================================================================

class PMSHierarchySearcher:
    """
    کلاس جستجوی سلسله‌مراتبی در ساختار outline PMS
    """

    def __init__(self, config: PMSConfig):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config

    @staticmethod
    def get_outline_level(row) -> int:
        """
        دریافت outline level یک سطر

        Args:
            row: row dimension از openpyxl

        Returns:
            سطح outline (0 اگر تعریف نشده)
        """
        return row.outlineLevel if row.outlineLevel else 0

    def find_items(self, ws: Worksheet, search_path: List[Tuple[int, str]],
                   target_level: int) -> List[Dict]:
        """
        جستجوی سلسله‌مراتبی برای یافتن آیتم‌های سطح هدف

        Args:
            ws: worksheet
            search_path: مسیر جستجو [(level, text), ...]
            target_level: سطح هدف برای استخراج آیتم‌ها

        Returns:
            لیست آیتم‌ها [{'row': ..., 'level': ..., 'text': ...}]
        """
        # مرحله 1: پیدا کردن مسیر کامل
        parent_row, search_start = self._find_parent_section(ws, search_path)

        if parent_row is None:
            return []

        # مرحله 2: استخراج آیتم‌های سطح هدف
        return self._extract_target_items(ws, parent_row, search_start, target_level)

    def _find_parent_section(self, ws: Worksheet,
                             search_path: List[Tuple[int, str]]) -> Tuple[Optional[int], Optional[int]]:
        """
        پیدا کردن بخش والد بر اساس مسیر سلسله‌مراتبی

        Args:
            ws: worksheet
            search_path: مسیر جستجو

        Returns:
            (شماره سطر والد، شماره سطر شروع جستجو) یا (None, None)
        """
        current_idx = 0
        parent_row = None

        for row_idx in range(1, ws.max_row + 1):
            if current_idx >= len(search_path):
                return parent_row, row_idx

            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)
            cell_value = ws.cell(row_idx, self.config.PMS.TEXT_COL).value

            if cell_value is None:
                continue

            cell_text = str(cell_value).strip()
            target_level, target_text = search_path[current_idx]

            if level == target_level and target_text in cell_text:
                parent_row = row_idx
                current_idx += 1

        return (None, None) if current_idx < len(search_path) else (parent_row, parent_row + 1)

    def _extract_target_items(self, ws: Worksheet, parent_row: int,
                              start_row: int, target_level: int) -> List[Dict]:
        """
        استخراج آیتم‌های سطح هدف از زیر بخش والد

        Args:
            ws: worksheet
            parent_row: شماره سطر والد
            start_row: شماره سطر شروع
            target_level: سطح هدف

        Returns:
            لیست آیتم‌ها
        """
        parent_level = self.get_outline_level(ws.row_dimensions[parent_row])
        found_items = []

        for row_idx in range(start_row, ws.max_row + 1):
            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)
            cell_value = ws.cell(row_idx, self.config.PMS.TEXT_COL).value

            if cell_value is None:
                continue

            # اگر به سطح مساوی یا کمتر از والد رسیدیم، توقف
            if level <= parent_level:
                break

            # اگر سطح هدف بود، ذخیره کن
            if level == target_level:
                found_items.append({
                    'row': row_idx,
                    'level': level,
                    'text': str(cell_value).strip()
                })

        return found_items

    def find_last_level5_in_section(self, file_path: str, sheet_name: str,
                                    mohor_num: int) -> Optional[int]:
        """
        پیدا کردن آخرین Level 5 در بخش "بلاست و اماده سازی..."

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت
            mohor_num: شماره محور

        Returns:
            شماره آخرین سطر Level 5 یا None
        """
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]

        search_path = self.config.Hierarchy.get_search_path(mohor_num)
        parent_row, search_start = self._find_parent_section(ws, search_path)

        if parent_row is None:
            wb.close()
            return None

        # پیدا کردن آخرین Level 5
        section_level = self.get_outline_level(ws.row_dimensions[parent_row])
        last_level5 = None

        for row_idx in range(search_start, ws.max_row + 1):
            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)

            if level <= section_level:
                break

            if level == self.config.Hierarchy.TARGET_LEVEL:
                last_level5 = row_idx

        wb.close()
        return last_level5

# ================================================================================
# 📄 ماژول استخراج آیتم‌های PNT
# ================================================================================

class PNTItemExtractor:
    """
    کلاس استخراج آیتم‌ها از فایل PNT-G
    """

    def __init__(self, config: PMSConfig, log_callback=None):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
            log_callback: تابع callback برای ارسال لاگ‌ها
        """
        self.config = config
        self.log_callback = log_callback or print
        self.normalizer = TextNormalizer()
        self.axis_extractor = AxisExtractor(config)

    def extract_all_items(self, file_path: str, sheet_name: str) -> Tuple[Dict[int, List[Dict]], List[Dict], Any]:
        """
        استخراج تمام آیتم‌های PNT-G با شناسایی محور

        Args:
            file_path: مسیر فایل PNT-G
            sheet_name: نام شیت

        Returns:
            (items_by_axis, unidentified_items, g2_value)
        """
        self.log_callback(f"\n📂 بارگذاری {file_path}...", "info")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        # خواندن و پاکسازی مقدار G2
        g2_raw = ws[self.config.PNT.G2_CELL].value
        g2_value = self.normalizer.clean_g2_value(g2_raw)

        self.log_callback(f"✅ مقدار G2 خام: {g2_raw}", "info")
        self.log_callback(f"✅ مقدار G2 پاکسازی شده: {g2_value}", "success")

        # استخراج آیتم‌ها
        items_by_axis = {}
        unidentified = []

        for row in range(self.config.PNT.ROW_START, self.config.PNT.ROW_END):
            item_data = self._extract_row_data(ws, row)

            if item_data is None:
                continue

            if item_data['axis'] is None:
                unidentified.append({
                    'row': row,
                    'item': item_data['single_line']
                })
                continue

            # گروه‌بندی بر اساس محور
            axis_num = item_data['axis']
            if axis_num not in items_by_axis:
                items_by_axis[axis_num] = []

            items_by_axis[axis_num].append(item_data)

        wb.close()

        total_items = sum(len(items) for items in items_by_axis.values())
        self.log_callback(f"✅ {total_items} آیتم استخراج شد از {len(items_by_axis)} محور", "success")

        if unidentified:
            self.log_callback(f"\n⚠️  {len(unidentified)} آیتم بدون محور:", "warning")
            for item in unidentified[:5]:
                self.log_callback(f"   ❌ سطر {item['row']}: {item['item']}", "warning")
            if len(unidentified) > 5:
                self.log_callback(f"   ... و {len(unidentified) - 5} مورد دیگر", "warning")

        return items_by_axis, unidentified, g2_value

    def _extract_row_data(self, ws: Worksheet, row: int) -> Optional[Dict]:
        """
        استخراج داده‌های یک سطر PNT-G

        Args:
            ws: worksheet
            row: شماره سطر

        Returns:
            دیکشنری داده‌های سطر یا None
        """
        item_value = ws.cell(row, self.config.PNT.ITEM_COL).value

        if not item_value:
            return None

        # استخراج محور
        axis_num = self.axis_extractor.extract_from_row(ws, row)

        # پردازش متن
        original_text = str(item_value).strip()
        single_line = self.normalizer.multiline_to_single(original_text)
        normalized = self.normalizer.normalize_standard(single_line)

        if not normalized:
            return None

        # خواندن مقادیر
        quantity = ws.cell(row, self.config.PNT.QUANTITY_COL).value
        m_value = ws.cell(row, self.config.PNT.M_VALUE_COL).value

        return {
            'pnt_row': row,
            'quantity': int(quantity) if quantity else 0,
            'm_value': m_value,
            'original': original_text,
            'single_line': single_line,
            'normalized': normalized,
            'axis': axis_num
        }

# ================================================================================
# 🔄 ماژول برنامه‌ریزی به‌روزرسانی
# ================================================================================

class UpdatePlanner:
    """
    کلاس برنامه‌ریزی و تطابق آیتم‌های PNT با PMS
    """

    def __init__(self, config: PMSConfig, log_callback=None):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
            log_callback: تابع callback برای ارسال لاگ‌ها
        """
        self.config = config
        self.log_callback = log_callback or print
        self.hierarchy_searcher = PMSHierarchySearcher(config)

    def plan_updates(self, pms_file: str, pms_sheet: str,
                     item_locations: Dict, items_by_axis: Dict[int, List[Dict]],
                     g2_value: Any) -> Tuple[List[Dict], List[Dict], List[Dict]]:
        """
        برنامه‌ریزی به‌روزرسانی‌ها با تطابق آیتم‌ها

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت PMS
            item_locations: دیکشنری موقعیت آیتم‌ها در PMS
            items_by_axis: آیتم‌های PNT گروه‌بندی شده بر اساس محور
            g2_value: مقدار سلول G2 از PNT

        Returns:
            (updates, not_found, warnings)
        """
        self.log_callback("\n🔗 تطابق آیتم‌ها با PMS...", "info")

        updates = []
        not_found = []
        warnings = []
        found_existing = 0
        found_new = 0

        for mohor_num, items in items_by_axis.items():
            mohor_name = f"محور {mohor_num}"

            for pnt_data in items:
                result = self._match_item(
                    pms_file, pms_sheet, mohor_name, mohor_num,
                    pnt_data, item_locations, g2_value
                )

                if result['status'] == 'existing':
                    found_existing += 1
                    updates.append(result['update'])
                    if result.get('warning'):
                        warnings.append(result['warning'])

                elif result['status'] == 'new':
                    found_new += 1
                    updates.append(result['update'])

                elif result['status'] == 'not_found':
                    not_found.append(result['error'])

        # گزارش تطابق
        self.log_callback(f"\n✅ آیتم‌های موجود: {found_existing}", "success")
        self.log_callback(f"🆕 آیتم‌های جدید: {found_new}", "success")
        self.log_callback(f"❌ آیتم‌های قابل درج نیستند: {len(not_found)}", "error" if not_found else "info")

        if warnings:
            self.log_callback(f"\n⚠️  {len(warnings)} آیتم نیاز به درج سطر دارند:", "warning")
            for w in warnings[:5]:
                self.log_callback(f"   - {w['item']} ({w['mohor']}): کمبود {w['deficit']} سطر", "warning")
            if len(warnings) > 5:
                self.log_callback(f"   ... و {len(warnings) - 5} مورد دیگر", "warning")

        if not_found:
            self.log_callback(f"\n❌ آیتم‌های قابل درج نیستند:", "error")
            for item in not_found[:5]:
                self.log_callback(f"   - {item['item']} ({item['mohor']}): {item['reason']}", "error")
            if len(not_found) > 5:
                self.log_callback(f"   ... و {len(not_found) - 5} مورد دیگر", "error")

        return updates, not_found, warnings

    def _match_item(self, pms_file: str, pms_sheet: str, mohor_name: str,
                    mohor_num: int, pnt_data: Dict, item_locations: Dict,
                    g2_value: Any) -> Dict:
        """
        تطابق یک آیتم PNT با PMS

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت
            mohor_name: نام محور
            mohor_num: شماره محور
            pnt_data: داده‌های آیتم از PNT
            item_locations: دیکشنری موقعیت‌ها
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه تطابق
        """
        normalized_item = pnt_data['normalized']

        # بررسی وجود در PMS
        if normalized_item in item_locations:
            mohor_locations = [
                loc for loc in item_locations[normalized_item]
                if loc['mohor'] == mohor_name
            ]

            if mohor_locations:
                # آیتم موجود است
                return self._create_existing_update(
                    mohor_name, pnt_data, mohor_locations, g2_value
                )

        # آیتم جدید - باید درج شود
        return self._create_new_update(
            pms_file, pms_sheet, mohor_name, mohor_num, pnt_data, g2_value
        )

    def _create_existing_update(self, mohor_name: str, pnt_data: Dict,
                                locations: List[Dict], g2_value: Any) -> Dict:
        """
        ایجاد به‌روزرسانی برای آیتم موجود

        Args:
            mohor_name: نام محور
            pnt_data: داده‌های PNT
            locations: لیست موقعیت‌های آیتم در PMS
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه
        """
        needed_quantity = pnt_data['quantity']
        current_quantity = len(locations)

        result = {
            'status': 'existing',
            'update': {
                'mohor': mohor_name,
                'item_text': pnt_data['single_line'],
                'existing_rows': [loc['row'] for loc in locations],
                'needed_quantity': needed_quantity,
                'a_value': pnt_data['single_line'],
                'e_value': g2_value,
                'n_value': pnt_data['m_value'],
                'is_new_item': False
            }
        }

        # بررسی کمبود
        if current_quantity < needed_quantity:
            deficit = needed_quantity - current_quantity
            result['warning'] = {
                'item': pnt_data['single_line'],
                'mohor': mohor_name,
                'needed': needed_quantity,
                'available': current_quantity,
                'deficit': deficit
            }

        return result

    def _create_new_update(self, pms_file: str, pms_sheet: str,
                           mohor_name: str, mohor_num: int,
                           pnt_data: Dict, g2_value: Any) -> Dict:
        """
        ایجاد به‌روزرسانی برای آیتم جدید

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت
            mohor_name: نام محور
            mohor_num: شماره محور
            pnt_data: داده‌های PNT
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه
        """
        last_level5 = self.hierarchy_searcher.find_last_level5_in_section(
            pms_file, pms_sheet, mohor_num
        )

        if last_level5:
            return {
                'status': 'new',
                'update': {
                    'mohor': mohor_name,
                    'item_text': f"🆕 {pnt_data['single_line']}",
                    'existing_rows': [last_level5],
                    'needed_quantity': pnt_data['quantity'],
                    'a_value': pnt_data['single_line'],
                    'e_value': g2_value,
                    'n_value': pnt_data['m_value'],
                    'g_value': pnt_data['m_value'],
                    'is_new_item': True
                }
            }
        else:
            return {
                'status': 'not_found',
                'error': {
                    'item': pnt_data['single_line'],
                    'mohor': mohor_name,
                    'reason': 'محور یا Level 5 در PMS یافت نشد'
                }
            }

# ================================================================================
# 🔄 ماژول به‌روزرسانی با COM
# ================================================================================

class COMUpdater:
    """
    کلاس به‌روزرسانی فایل Excel با استفاده از win32com
    """

    def __init__(self, config: PMSConfig, log_callback=None):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
            log_callback: تابع callback برای ارسال لاگ‌ها
        """
        self.config = config
        self.log_callback = log_callback or print

    def update_file(self, file_path: str, sheet_name: str, updates: List[Dict]):
        """
        به‌روزرسانی فایل PMS

        Args:
            file_path: مسیر فایل
            sheet_name: نام شیت
            updates: لیست به‌روزرسانی‌ها
        """
        self.log_callback("\n" + "=" * 50, "info")
        self.log_callback("🔧 باز کردن Excel با COM...", "info")
        self.log_callback("=" * 50, "info")

        abs_path = os.path.abspath(file_path)
        self.log_callback(f"📂 مسیر فایل: {abs_path}", "info")

        if not os.path.exists(abs_path):
            raise FileNotFoundError(f"❌ فایل یافت نشد: {abs_path}")

        # باز کردن فایل با openpyxl برای بررسی ستون E
        self.log_callback("🔍 بررسی وضعیت ستون E...", "info")
        wb_check = openpyxl.load_workbook(abs_path, data_only=True)
        ws_check = wb_check[sheet_name]

        xl = None
        wb = None

        try:
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False

            self.log_callback(f"🔓 در حال باز کردن Workbook...", "info")
            wb = xl.Workbooks.Open(abs_path)

            if wb is None:
                raise RuntimeError("❌ Workbook باز نشد")

            ws = wb.Worksheets(sheet_name)
            self.log_callback(f"✅ شیت '{sheet_name}' یافت شد", "success")

            # پردازش به‌روزرسانی‌ها
            stats = self._process_updates(ws, ws_check, updates)

            # ذخیره
            self.log_callback(f"\n💾 ذخیره‌سازی...", "info")
            wb.Save()

            # گزارش نهایی
            self._print_stats(stats)

        except Exception as e:
            self.log_callback(f"\n❌ خطا: {e}", "error")
            import traceback
            traceback.print_exc()
            raise

        finally:
            try:
                wb_check.close()
                if wb:
                    wb.Close(SaveChanges=False)
                if xl:
                    xl.Quit()
            except:
                pass

    def _process_updates(self, ws_com, ws_check, updates: List[Dict]) -> Dict:
        """
        پردازش لیست به‌روزرسانی‌ها

        Args:
            ws_com: worksheet COM
            ws_check: worksheet openpyxl (برای چک کردن ستون E)
            updates: لیست به‌روزرسانی‌ها

        Returns:
            دیکشنری آمار
        """
        stats = {
            'inserted': 0,
            'updated': 0,
            'skipped': 0
        }

        for update in updates:
            if update.get('is_new_item'):
                stats['inserted'] += self._process_new_item(ws_com, update)
            else:
                result = self._process_existing_item(ws_com, ws_check, update)
                stats['inserted'] += result['inserted']
                stats['updated'] += result['updated']
                stats['skipped'] += result['skipped']

        return stats

    def _process_new_item(self, ws, update: Dict) -> int:
        """
        پردازش آیتم جدید (درج + کپی کامل از ردیف الگو + آپدیت A, E, G, N)

        Args:
            ws: worksheet COM
            update: دیکشنری به‌روزرسانی

        Returns:
            تعداد ردیف‌های درج شده
        """
        item_text = update['item_text']
        existing_rows = update['existing_rows']
        needed_quantity = update['needed_quantity']
        a_value = update['a_value']
        e_value = update['e_value']
        g_value = update['g_value']
        n_value = update['n_value']

        self.log_callback(f"\n📝 {item_text}", "info")
        self.log_callback(f"   🆕 آیتم جدید - درج {needed_quantity} ردیف", "info")

        # ردیف الگو = آخرین ردیف Level 5 پیدا شده
        template_row = existing_rows[-1]
        inserted_count = 0

        for i in range(needed_quantity):
            new_row = template_row + i + 1
            ws.Rows(new_row).Insert()

            # ✅ مرحله 1: کپی کامل ردیف (شامل تمام مقادیر و فرمول‌ها)
            self._copy_row_with_values(ws, template_row, new_row)

            # ✅ مرحله 2: بازنویسی فقط ستون‌های A, E, G, N
            ws.Cells(new_row, self.config.PMS.TEXT_COL).Value = a_value
            ws.Cells(new_row, self.config.PMS.DATE_COL).Value = e_value
            ws.Cells(new_row, self.config.PMS.G_COL).Value = g_value
            ws.Cells(new_row, self.config.PMS.N_COL).Value = n_value

            existing_rows.append(new_row)
            inserted_count += 1

        self.log_callback(f"   ✅ {needed_quantity} ردیف درج شد", "success")
        return inserted_count

    def _process_existing_item(self, ws_com, ws_check, update: Dict) -> Dict:
        """
        پردازش آیتم موجود (فیلترینگ E + درج در صورت کمبود + آپدیت A, E, N)

        Args:
            ws_com: worksheet COM
            ws_check: worksheet openpyxl
            update: دیکشنری به‌روزرسانی

        Returns:
            دیکشنری آمار {'inserted': ..., 'updated': ..., 'skipped': ...}
        """
        item_text = update['item_text']
        existing_rows = update['existing_rows']
        needed_quantity = update['needed_quantity']
        a_value = update['a_value']
        e_value = update['e_value']
        n_value = update['n_value']

        self.log_callback(f"\n📝 {item_text}", "info")

        # فیلتر ردیف‌های خالی و پُر
        filled_rows = []
        empty_rows = []

        for row in existing_rows:
            e_cell_value = ws_check.cell(row, self.config.PMS.DATE_COL).value
            if e_cell_value is None or str(e_cell_value).strip() == "":
                empty_rows.append(row)
            else:
                filled_rows.append(row)

        stats = {'inserted': 0, 'updated': 0, 'skipped': len(filled_rows)}

        if filled_rows:
            self.log_callback(f"   ⏭️  {len(filled_rows)} ردیف پُر (نادیده گرفته شد)", "warning")

        self.log_callback(f"   📊 موجود: {len(existing_rows)} | خالی: {len(empty_rows)} | نیاز: {needed_quantity}", "info")

        # محاسبه کمبود
        deficit = max(0, needed_quantity - len(empty_rows))

        if deficit > 0:
            self.log_callback(f"   ➕ درج {deficit} ردیف جدید...", "info")
            last_row = existing_rows[-1]

            for i in range(deficit):
                new_row = last_row + i + 1
                ws_com.Rows(new_row).Insert()

                # ✅ کپی کامل ردیف (با مقادیر)
                self._copy_row_with_values(ws_com, last_row, new_row)

                # آپدیت فقط A, E, N (G دست نخورده)
                ws_com.Cells(new_row, self.config.PMS.TEXT_COL).Value = a_value
                ws_com.Cells(new_row, self.config.PMS.DATE_COL).Value = e_value
                ws_com.Cells(new_row, self.config.PMS.N_COL).Value = n_value

                empty_rows.append(new_row)
                existing_rows.append(new_row)
                stats['inserted'] += 1

        # آپدیت ردیف‌های خالی
        rows_to_update = empty_rows[:needed_quantity]

        for row in rows_to_update:
            ws_com.Cells(row, self.config.PMS.TEXT_COL).Value = a_value
            ws_com.Cells(row, self.config.PMS.DATE_COL).Value = e_value
            ws_com.Cells(row, self.config.PMS.N_COL).Value = n_value
            stats['updated'] += 1

        self.log_callback(f"   ✅ {len(rows_to_update)} ردیف آپدیت شد", "success")
        self.log_callback(f"   📍 ردیف‌ها: {', '.join(map(str, rows_to_update))}", "info")

        return stats

    def _copy_row_with_values(self, ws, source_row: int, target_row: int):
        """
        کپی کامل یک ردیف (محتوا + فرمول + استایل + outline)

        Args:
            ws: worksheet COM
            source_row: شماره ردیف مبدا (الگو)
            target_row: شماره ردیف مقصد (جدید)
        """
        try:
            source_range = ws.Rows(source_row)
            target_range = ws.Rows(target_row)

            # کپی کامل (All = محتوا + فرمت + فرمول)
            source_range.Copy()
            target_range.PasteSpecial(Paste=-4104)  # xlPasteAll

            # تنظیم outline level
            target_range.OutlineLevel = source_range.OutlineLevel

        except Exception as e:
            self.log_callback(f"⚠️  خطا در کپی ردیف {source_row}: {e}", "warning")
            raise

    def _print_stats(self, stats: Dict):
        """
        چاپ آمار نهایی

        Args:
            stats: دیکشنری آمار
        """
        self.log_callback(f"\n" + "=" * 50, "info")
        self.log_callback(f"✅ خلاصه:", "success")
        self.log_callback(f"   🆕 ردیف‌های درج شده: {stats['inserted']}", "success")
        self.log_callback(f"   🔄 ردیف‌های آپدیت شده: {stats['updated']}", "success")
        self.log_callback(f"   ⏭️  ردیف‌های نادیده گرفته شده (E پُر): {stats['skipped']}", "info")
        self.log_callback("=" * 50, "info")


# ================================================================================
# 🚀 هماهنگ‌کننده اصلی
# ================================================================================

# ================================================================================
# 🚀 هماهنگ‌کننده اصلی
# ================================================================================

class PMSUpdateOrchestrator:
    """
    کلاس هماهنگی کل فرآیند به‌روزرسانی PMS از PNT-G
    """

    def __init__(self, config: PMSConfig, log_callback=None):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
            log_callback: تابع callback برای ارسال لاگ‌ها (msg, type)
        """
        self.config = config
        self.log_callback = log_callback or self._default_log
        self.cache_manager = PMSCacheManager(config.CACHE_FILE, self.log_callback)
        self.structure_reader = PMSStructureReader(config, self.log_callback)
        self.pnt_extractor = PNTItemExtractor(config, self.log_callback)
        self.update_planner = UpdatePlanner(config, self.log_callback)
        self.com_updater = COMUpdater(config, self.log_callback)

    @staticmethod
    def _default_log(msg: str, msg_type: str = 'info'):
        """لاگ پیش‌فرض (برای استفاده در خط فرمان)"""
        print(msg)

    def run(self):
        """
        اجرای کامل فرآیند به‌روزرسانی

        Returns:
            دیکشنری نتایج
        """
        self.log_callback("=" * 50, "info")
        self.log_callback("🚀 شروع فرآیند کامل به‌روزرسانی PMS از PNT-G", "info")
        self.log_callback("=" * 50, "info")

        # مرحله 1: بارگذاری ساختار PMS (با Cache)
        self.log_callback("\n📥 بارگذاری ساختار PMS...", "info")
        item_locations = self._load_pms_structure()
        self.log_callback(f"✅ {len(item_locations)} آیتم منحصر به فرد", "success")

        # مرحله 2: استخراج آیتم‌های PNT-G
        self.log_callback(f"\n🔍 استخراج آیتم‌های PNT-G...", "info")
        items_by_axis, unidentified, g2_value = self.pnt_extractor.extract_all_items(
            self.config.PNT_FILE,
            self.config.PNT_SHEET
        )

        # مرحله 3: تطابق و برنامه‌ریزی به‌روزرسانی‌ها
        updates, not_found, warnings = self.update_planner.plan_updates(
            self.config.PMS_FILE,
            self.config.PMS_SHEET,
            item_locations,
            items_by_axis,
            g2_value
        )

        # مرحله 4: اجرای به‌روزرسانی
        if updates:
            self.com_updater.update_file(
                self.config.PMS_FILE,
                self.config.PMS_SHEET,
                updates
            )
        else:
            self.log_callback("\n⚠️  هیچ به‌روزرسانی‌ای برای اجرا وجود ندارد!", "warning")

        # گزارش نهایی
        self._print_final_report(updates, not_found, warnings, unidentified)

        # ✅ برگرداندن نتایج کامل (شامل لیست‌ها)
        return {
            'processed': len(updates),
            'not_found': len(not_found),
            'warnings': len(warnings),
            'unidentified_axis': len(unidentified),
            'updates': updates,
            'not_found_list': not_found,
            'warnings_list': warnings,
            'unidentified_list': unidentified,
            'dry_run': False
        }

    def _load_pms_structure(self) -> Dict:
        """
        بارگذاری ساختار PMS با استفاده از Cache

        Returns:
            دیکشنری موقعیت آیتم‌ها
        """
        # تلاش برای بارگذاری از Cache
        if self.config.USE_CACHE:
            cached_data = self.cache_manager.load_cache(
                self.config.PMS_FILE,
                self.config.PMS_SHEET
            )

            if cached_data:
                self.log_callback("✅ استفاده از Cache (فایل تغییر نکرده)", "success")
                return cached_data

        # استخراج از فایل
        item_locations = self.structure_reader.extract_all_items(
            self.config.PMS_FILE,
            self.config.PMS_SHEET
        )

        # ذخیره Cache
        if self.config.USE_CACHE:
            self.cache_manager.save_cache(
                self.config.PMS_FILE,
                self.config.PMS_SHEET,
                item_locations
            )

        return item_locations

    def _print_final_report(self, updates, not_found, warnings, unidentified):
        """
        چاپ گزارش نهایی

        Args:
            updates: لیست به‌روزرسانی‌ها
            not_found: لیست آیتم‌های پیدا نشده
            warnings: لیست هشدارها
            unidentified: لیست آیتم‌های بدون محور
        """
        self.log_callback("\n" + "=" * 50, "info")
        self.log_callback("🏁 پایان عملیات", "success")
        self.log_callback("=" * 50, "info")
        self.log_callback(f"📊 خلاصه گزارش:", "info")

        existing_count = sum(1 for u in updates if not u.get('is_new_item'))
        new_count = sum(1 for u in updates if u.get('is_new_item'))

        self.log_callback(f"   ✅ آیتم‌های موجود پردازش شده: {existing_count}", "success")
        self.log_callback(f"   🆕 آیتم‌های جدید درج شده: {new_count}", "success")
        self.log_callback(f"   ❌ آیتم‌های ناموفق: {len(not_found)}", "error" if not_found else "info")
        self.log_callback(f"   ⚠️  آیتم‌های دارای کمبود: {len(warnings)}", "warning" if warnings else "info")

        if unidentified:
            self.log_callback(f"   🔍 آیتم‌های بدون محور: {len(unidentified)}", "warning")

        self.log_callback("=" * 50, "info")

# ================================================================================
# 🎯 نقطه ورود برنامه
# ================================================================================

def main():
    """
    تابع اصلی برنامه
    """
    try:
        # ایجاد orchestrator با تنظیمات پیش‌فرض
        orchestrator = PMSUpdateOrchestrator(PMSConfig())

        # اجرای فرآیند
        results = orchestrator.run()

        print("\n✅ عملیات با موفقیت انجام شد!")

        return results

    except FileNotFoundError as e:
        print(f"\n❌ خطای فایل: {e}")
        return None
    except Exception as e:
        print(f"\n❌ خطای کلی: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    main()
