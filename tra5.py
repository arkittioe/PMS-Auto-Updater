import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re
import json
import os
from datetime import datetime
import win32com.client
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any


# ================================================================================
# 🎛️ تنظیمات مرکزی
# ================================================================================

class PMSConfig:
    """
    کلاس تنظیمات مرکزی برنامه
    تمام مسیرها، نام‌ها و mapping ستون‌ها در اینجا تعریف می‌شوند
    """

    # ═══════════════════════════════════════════════════════════════════════════
    # فایل‌ها و شیت‌ها
    # ═══════════════════════════════════════════════════════════════════════════
    PMS_FILE = "PMS-paint REV-03H.xlsx"
    PMS_SHEET = "1404.01.22"
    PNT_FILE = "PNT-G.xlsx"
    PNT_SHEET = " المان PNT-G-130"
    CACHE_FILE = "pms_cache.json"

    # ═══════════════════════════════════════════════════════════════════════════
    # ستون‌های PNT-G
    # ═══════════════════════════════════════════════════════════════════════════
    class PNT:
        """تنظیمات مربوط به فایل PNT-G"""
        ITEM_COL = 3  # C - متن آیتم
        QUANTITY_COL = 9  # I - تعداد
        M_VALUE_COL = 13  # M - مقدار برای ستون N در PMS
        G2_CELL = "G2"  # سلول تاریخ
        ROW_START = 7  # شروع محدوده داده
        ROW_END = 31  # پایان محدوده داده
        AXIS_SEARCH_COLS = [3, 4, 5]  # C, D, E - ستون‌های جستجوی محور

    # ═══════════════════════════════════════════════════════════════════════════
    # ستون‌های PMS
    # ═══════════════════════════════════════════════════════════════════════════
    class PMS:
        """تنظیمات مربوط به فایل PMS"""
        TEXT_COL = 1  # A - متن آیتم
        DATE_COL = 5  # E - تاریخ (از G2 فایل PNT)
        G_COL = 7  # G - مقدار g (فقط برای آیتم جدید)
        N_COL = 14  # N - مقدار از ستون M فایل PNT

    # ═══════════════════════════════════════════════════════════════════════════
    # محدوده محورها
    # ═══════════════════════════════════════════════════════════════════════════
    AXIS_RANGE_START = 19  # شروع محور
    AXIS_RANGE_END = 46  # پایان محور (exclusive)

    # ═══════════════════════════════════════════════════════════════════════════
    # مسیر سلسله‌مراتبی برای جستجو در PMS
    # ═══════════════════════════════════════════════════════════════════════════
    class Hierarchy:
        """تنظیمات مسیر سلسله‌مراتبی"""
        LEVEL_1_PATTERN = "محور"  # الگوی محور در Level 1
        LEVEL_3_TEXT = "GLASS FLAKE"
        LEVEL_4_TEXT = "بلاست و اماده سازی سطح  و اعمال رنگ  لایه دوم"
        TARGET_LEVEL = 5  # سطح هدف برای آیتم‌ها

        @classmethod
        def get_search_path(cls, mohor_num: int) -> List[Tuple[int, str]]:
            """
            ساخت مسیر جستجوی سلسله‌مراتبی برای محور مشخص

            Args:
                mohor_num: شماره محور (19-45)

            Returns:
                لیست تاپل‌های (level, text)
            """
            return [
                (1, f"{cls.LEVEL_1_PATTERN} {mohor_num}"),
                (3, cls.LEVEL_3_TEXT),
                (4, cls.LEVEL_4_TEXT)
            ]


# ================================================================================
# 🔧 ماژول نرمال‌سازی متن
# ================================================================================

class TextNormalizer:
    """
    کلاس نرمال‌سازی متن
    شامل متدهای مختلف برای نرمال‌سازی متن فارسی و انگلیسی
    """

    @staticmethod
    def clean_g2_value(text: Optional[str]) -> str:
        """
        پاکسازی مقدار G2 برای استفاده در PMS
        - حذف عبارت "شماره صورتمجلس"
        - حذف فاصله‌های اضافی

        Args:
            text: متن ورودی از سلول G2

        Returns:
            متن پاکسازی شده
        """
        if text is None:
            return ""

        text = str(text).strip()

        # حذف عبارت "شماره صورتمجلس" (با حروف مختلف ی/ی و ک/ک)
        text = re.sub(r'شماره\s*صورت\s*مجلس', '', text, flags=re.IGNORECASE)
        text = re.sub(r'شماره\s*صورتمجلس', '', text, flags=re.IGNORECASE)

        # حذف فاصله‌های اضافی
        text = ' '.join(text.split())

        return text.strip()

    @staticmethod
    def normalize_standard(text: Optional[str]) -> str:
        """
        نرمال‌سازی استاندارد برای متن فارسی

        Args:
            text: متن ورودی

        Returns:
            متن نرمال‌شده (lowercase, بدون فاصله)
        """
        if text is None:
            return ""

        text = str(text).strip()
        text = text.replace('ی', 'ی').replace('ک', 'ک')
        text = re.sub(r'\s+', '', text)
        text = text.lower()

        return text

    @staticmethod
    def normalize_pnt_axis(text: Optional[str]) -> str:
        """
        نرمال‌سازی پیشرفته برای استخراج محور از PNT-G
        - تبدیل \\n به فاصله
        - تبدیل به uppercase
        - حذف فاصله‌ها و خط‌تیره

        Args:
            text: متن ورودی

        Returns:
            متن نرمال‌شده (UPPERCASE, بدون فاصله و خط‌تیره)
        """
        if text is None:
            return ""

        text = str(text)
        # تبدیل خطوط به فاصله
        text = text.replace('\n', ' ').replace('\r', ' ')
        # تبدیل به حروف بزرگ
        text = text.upper()
        # حذف فاصله‌ها و خط‌تیره
        text = text.replace(' ', '').replace('-', '')

        return text

    @staticmethod
    def multiline_to_single(text: Optional[str]) -> str:
        """
        تبدیل متن چندخطی به تک‌خطی
        - تبدیل \\n و \\r به فاصله
        - حذف فاصه‌های اضافی

        Args:
            text: متن چندخطی

        Returns:
            متن تک‌خطی
        """
        if text is None:
            return ""

        text = str(text)
        # تبدیل خطوط به فاصله
        text = text.replace('\n', ' ').replace('\r', ' ')
        # حذف فاصله‌های متوالی
        text = ' '.join(text.split())

        return text.strip()


# ================================================================================
# 🔧 ماژول استخراج محور
# ================================================================================

class AxisExtractor:
    """
    کلاس استخراج شماره محور از سطرهای PNT-G
    با دو اولویت جستجو: AXIS[19-45] و سپس S[19-45]
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.normalizer = TextNormalizer()

    def extract_from_row(self, ws: Worksheet, row_idx: int) -> Optional[int]:
        """
        استخراج شماره محور از یک سطر PNT-G

        جستجو در ستون‌های مشخص شده با دو اولویت:
        1. AXIS[19-45]
        2. S[19-45]

        Args:
            ws: worksheet PNT-G
            row_idx: شماره سطر

        Returns:
            شماره محور (19-45) یا None
        """
        # اولویت 1: جستجوی AXIS[19-45]
        mohor = self._search_pattern(ws, row_idx, "AXIS")
        if mohor:
            return mohor

        # اولویت 2: جستجوی S[19-45]
        mohor = self._search_pattern(ws, row_idx, "S")
        return mohor

    def _search_pattern(self, ws: Worksheet, row_idx: int, pattern_prefix: str) -> Optional[int]:
        """
        جستجوی الگو در ستون‌های مشخص

        Args:
            ws: worksheet
            row_idx: شماره سطر
            pattern_prefix: پیشوند الگو (AXIS یا S)

        Returns:
            شماره محور یا None
        """
        for col in self.config.PNT.AXIS_SEARCH_COLS:
            cell_value = ws.cell(row_idx, col).value
            if not cell_value:
                continue

            normalized = self.normalizer.normalize_pnt_axis(cell_value)

            for mohor_num in range(self.config.AXIS_RANGE_START,
                                   self.config.AXIS_RANGE_END):
                pattern = f"{pattern_prefix}{mohor_num}"
                if pattern in normalized:
                    return mohor_num

        return None


# ================================================================================
# 💾 ماژول مدیریت Cache
# ================================================================================

class PMSCacheManager:
    """
    کلاس مدیریت cache ساختار PMS
    برای جلوگیری از استخراج مکرر ساختار فایل
    """

    def __init__(self, cache_file: str):
        """
        مقداردهی اولیه

        Args:
            cache_file: مسیر فایل cache
        """
        self.cache_file = cache_file

    @staticmethod
    def get_file_hash(file_path: str) -> str:
        """
        محاسبه hash ساده بر اساس modified time و size

        Args:
            file_path: مسیر فایل

        Returns:
            hash string
        """
        stat = os.stat(file_path)
        return f"{stat.st_mtime}_{stat.st_size}"

    def load_cache(self, file_path: str, sheet_name: str) -> Optional[Dict]:
        """
        بارگذاری cache در صورت معتبر بودن

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت

        Returns:
            دیکشنری cache یا None
        """
        if not os.path.exists(self.cache_file):
            return None

        try:
            with open(self.cache_file, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)

            current_hash = self.get_file_hash(file_path)

            if (cache_data.get('file_hash') == current_hash and
                    cache_data.get('sheet_name') == sheet_name):
                return cache_data['item_locations']
        except Exception as e:
            print(f"⚠️  خطا در خواندن Cache: {e}")

        return None

    def save_cache(self, file_path: str, sheet_name: str, item_locations: Dict):
        """
        ذخیره cache

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت
            item_locations: دیکشنری موقعیت آیتم‌ها
        """
        try:
            cache_dir = os.path.dirname(self.cache_file)
            if cache_dir and not os.path.exists(cache_dir):
                os.makedirs(cache_dir)

            cache_data = {
                'file_hash': self.get_file_hash(file_path),
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'item_locations': item_locations
            }

            with open(self.cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)

            print(f"💾 Cache ذخیره شد: {self.cache_file}")
        except Exception as e:
            print(f"⚠️  خطا در ذخیره Cache: {e}")


# ================================================================================
# 📊 ماژول خواندن ساختار PMS
# ================================================================================

class PMSStructureReader:
    """
    کلاس خواندن و استخراج ساختار سلسله‌مراتبی PMS
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.normalizer = TextNormalizer()
        self.hierarchy_searcher = PMSHierarchySearcher(config)

    def extract_all_items(self, file_path: str, sheet_name: str) -> Dict[str, List[Dict]]:
        """
        استخراج تمام آیتم‌های Level 5 از همه محورها

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت

        Returns:
            دیکشنری: {normalized_item: [{'mohor': ..., 'row': ..., 'level': ...}]}
        """
        print("🔄 در حال استخراج ساختار PMS...")

        # جستجو در تمام محورها
        mohor_results = self._search_all_mohors(file_path, sheet_name)

        # تبدیل به فرمت دیکشنری برای جستجوی سریع
        item_locations = {}
        for mohor_name, items in mohor_results.items():
            for item in items:
                normalized_text = self.normalizer.normalize_standard(item['text'])
                if normalized_text not in item_locations:
                    item_locations[normalized_text] = []
                item_locations[normalized_text].append({
                    'mohor': mohor_name,
                    'row': item['row'],
                    'level': item['level'],
                    'original_text': item['text']
                })

        return item_locations

    def _search_all_mohors(self, file_path: str, sheet_name: str) -> Dict[str, List[Dict]]:
        """
        جستجوی سلسله‌مراتبی در تمام محورها

        Args:
            file_path: مسیر فایل
            sheet_name: نام شیت

        Returns:
            دیکشنری: {mohor_name: [items]}
        """
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]

        all_results = {}

        for mohor_num in range(self.config.AXIS_RANGE_START,
                               self.config.AXIS_RANGE_END):
            mohor_name = f"محور {mohor_num}"
            search_path = self.config.Hierarchy.get_search_path(mohor_num)

            results = self.hierarchy_searcher.find_items(
                ws,
                search_path,
                self.config.Hierarchy.TARGET_LEVEL
            )

            if results:
                all_results[mohor_name] = results

        wb.close()
        return all_results


# ================================================================================
# 📊 ماژول جستجوی سلسله‌مراتبی
# ================================================================================

class PMSHierarchySearcher:
    """
    کلاس جستجوی سلسله‌مراتبی در ساختار outline PMS
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config

    @staticmethod
    def get_outline_level(row) -> int:
        """
        دریافت outline level یک سطر

        Args:
            row: row dimension از openpyxl

        Returns:
            سطح outline (0 اگر تعریف نشده)
        """
        return row.outlineLevel if row.outlineLevel else 0

    def find_items(self, ws: Worksheet, search_path: List[Tuple[int, str]],
                   target_level: int) -> List[Dict]:
        """
        جستجوی سلسله‌مراتبی برای یافتن آیتم‌های سطح هدف

        Args:
            ws: worksheet
            search_path: مسیر جستجو [(level, text), ...]
            target_level: سطح هدف برای استخراج آیتم‌ها

        Returns:
            لیست آیتم‌ها [{'row': ..., 'level': ..., 'text': ...}]
        """
        # مرحله 1: پیدا کردن مسیر کامل
        parent_row, search_start = self._find_parent_section(ws, search_path)

        if parent_row is None:
            return []

        # مرحله 2: استخراج آیتم‌های سطح هدف
        return self._extract_target_items(ws, parent_row, search_start, target_level)

    def _find_parent_section(self, ws: Worksheet,
                             search_path: List[Tuple[int, str]]) -> Tuple[Optional[int], Optional[int]]:
        """
        پیدا کردن بخش والد بر اساس مسیر سلسله‌مراتبی

        Args:
            ws: worksheet
            search_path: مسیر جستجو

        Returns:
            (شماره سطر والد، شماره سطر شروع جستجو) یا (None, None)
        """
        current_idx = 0
        parent_row = None

        for row_idx in range(1, ws.max_row + 1):
            if current_idx >= len(search_path):
                return parent_row, row_idx

            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)
            cell_value = ws.cell(row_idx, self.config.PMS.TEXT_COL).value

            if cell_value is None:
                continue

            cell_text = str(cell_value).strip()
            target_level, target_text = search_path[current_idx]

            if level == target_level and target_text in cell_text:
                parent_row = row_idx
                current_idx += 1

        return (None, None) if current_idx < len(search_path) else (parent_row, parent_row + 1)

    def _extract_target_items(self, ws: Worksheet, parent_row: int,
                              start_row: int, target_level: int) -> List[Dict]:
        """
        استخراج آیتم‌های سطح هدف از زیر بخش والد

        Args:
            ws: worksheet
            parent_row: شماره سطر والد
            start_row: شماره سطر شروع
            target_level: سطح هدف

        Returns:
            لیست آیتم‌ها
        """
        parent_level = self.get_outline_level(ws.row_dimensions[parent_row])
        found_items = []

        for row_idx in range(start_row, ws.max_row + 1):
            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)
            cell_value = ws.cell(row_idx, self.config.PMS.TEXT_COL).value

            if cell_value is None:
                continue

            # اگر به سطح مساوی یا کمتر از والد رسیدیم، توقف
            if level <= parent_level:
                break

            # اگر سطح هدف بود، ذخیره کن
            if level == target_level:
                found_items.append({
                    'row': row_idx,
                    'level': level,
                    'text': str(cell_value).strip()
                })

        return found_items

    def find_last_level5_in_section(self, file_path: str, sheet_name: str,
                                    mohor_num: int) -> Optional[int]:
        """
        پیدا کردن آخرین Level 5 در بخش "بلاست و اماده سازی..."

        Args:
            file_path: مسیر فایل PMS
            sheet_name: نام شیت
            mohor_num: شماره محور

        Returns:
            شماره آخرین سطر Level 5 یا None
        """
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]

        search_path = self.config.Hierarchy.get_search_path(mohor_num)
        parent_row, search_start = self._find_parent_section(ws, search_path)

        if parent_row is None:
            wb.close()
            return None

        # پیدا کردن آخرین Level 5
        section_level = self.get_outline_level(ws.row_dimensions[parent_row])
        last_level5 = None

        for row_idx in range(search_start, ws.max_row + 1):
            row = ws.row_dimensions[row_idx]
            level = self.get_outline_level(row)

            if level <= section_level:
                break

            if level == self.config.Hierarchy.TARGET_LEVEL:
                last_level5 = row_idx

        wb.close()
        return last_level5


# ================================================================================
# 📄 ماژول استخراج آیتم‌های PNT
# ================================================================================

class PNTItemExtractor:
    """
    کلاس استخراج آیتم‌ها از فایل PNT-G
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.normalizer = TextNormalizer()
        self.axis_extractor = AxisExtractor(config)

    def extract_all_items(self, file_path: str, sheet_name: str) -> Tuple[Dict[int, List[Dict]], List[Dict], Any]:
        """
        استخراج تمام آیتم‌های PNT-G با شناسایی محور

        Args:
            file_path: مسیر فایل PNT-G
            sheet_name: نام شیت

        Returns:
            (items_by_axis, unidentified_items, g2_value)
        """
        print(f"\n📂 بارگذاری {file_path}...")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        # خواندن و پاکسازی مقدار G2
        g2_raw = ws[self.config.PNT.G2_CELL].value
        g2_value = self.normalizer.clean_g2_value(g2_raw)

        print(f"✅ مقدار G2 خام: {g2_raw}")
        print(f"✅ مقدار G2 پاکسازی شده: {g2_value}")

        # استخراج آیتم‌ها
        items_by_axis = {}
        unidentified = []

        for row in range(self.config.PNT.ROW_START, self.config.PNT.ROW_END):
            item_data = self._extract_row_data(ws, row)

            if item_data is None:
                continue

            if item_data['axis'] is None:
                unidentified.append({
                    'row': row,
                    'item': item_data['single_line']
                })
                continue

            # گروه‌بندی بر اساس محور
            axis_num = item_data['axis']
            if axis_num not in items_by_axis:
                items_by_axis[axis_num] = []

            items_by_axis[axis_num].append(item_data)

        wb.close()

        total_items = sum(len(items) for items in items_by_axis.values())
        print(f"✅ {total_items} آیتم استخراج شد از {len(items_by_axis)} محور")

        if unidentified:
            print(f"\n⚠️  {len(unidentified)} آیتم بدون محور:")
            for item in unidentified[:5]:
                print(f"   ❌ سطر {item['row']}: {item['item']}")
            if len(unidentified) > 5:
                print(f"   ... و {len(unidentified) - 5} مورد دیگر")

        return items_by_axis, unidentified, g2_value

    def _extract_row_data(self, ws: Worksheet, row: int) -> Optional[Dict]:
        """
        استخراج داده‌های یک سطر PNT-G

        Args:
            ws: worksheet
            row: شماره سطر

        Returns:
            دیکشنری داده‌های سطر یا None
        """
        item_value = ws.cell(row, self.config.PNT.ITEM_COL).value

        if not item_value:
            return None

        # استخراج محور
        axis_num = self.axis_extractor.extract_from_row(ws, row)

        # پردازش متن
        original_text = str(item_value).strip()
        single_line = self.normalizer.multiline_to_single(original_text)
        normalized = self.normalizer.normalize_standard(single_line)

        if not normalized:
            return None

        # خواندن مقادیر
        quantity = ws.cell(row, self.config.PNT.QUANTITY_COL).value
        m_value = ws.cell(row, self.config.PNT.M_VALUE_COL).value

        return {
            'pnt_row': row,
            'quantity': int(quantity) if quantity else 0,
            'm_value': m_value,
            'original': original_text,
            'single_line': single_line,
            'normalized': normalized,
            'axis': axis_num
        }


# ================================================================================
# 🔄 ماژول برنامه‌ریزی به‌روزرسانی
# ================================================================================

class UpdatePlanner:
    """
    کلاس برنامه‌ریزی و تطابق آیتم‌های PNT با PMS
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.hierarchy_searcher = PMSHierarchySearcher(config)

    def plan_updates(self, pms_file: str, pms_sheet: str,
                     item_locations: Dict, items_by_axis: Dict[int, List[Dict]],
                     g2_value: Any) -> Tuple[List[Dict], List[Dict], List[Dict]]:
        """
        برنامه‌ریزی به‌روزرسانی‌ها با تطابق آیتم‌ها

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت PMS
            item_locations: دیکشنری موقعیت آیتم‌ها در PMS
            items_by_axis: آیتم‌های PNT گروه‌بندی شده بر اساس محور
            g2_value: مقدار سلول G2 از PNT

        Returns:
            (updates, not_found, warnings)
        """
        print("\n🔗 تطابق آیتم‌ها با PMS...")

        updates = []
        not_found = []
        warnings = []
        found_existing = 0
        found_new = 0

        for mohor_num, items in items_by_axis.items():
            mohor_name = f"محور {mohor_num}"

            for pnt_data in items:
                result = self._match_item(
                    pms_file, pms_sheet, mohor_name, mohor_num,
                    pnt_data, item_locations, g2_value
                )

                if result['status'] == 'existing':
                    found_existing += 1
                    updates.append(result['update'])
                    if result.get('warning'):
                        warnings.append(result['warning'])

                elif result['status'] == 'new':
                    found_new += 1
                    updates.append(result['update'])

                elif result['status'] == 'not_found':
                    not_found.append(result['error'])

        # گزارش تطابق
        print(f"\n✅ آیتم‌های موجود: {found_existing}")
        print(f"🆕 آیتم‌های جدید: {found_new}")
        print(f"❌ آیتم‌های قابل درج نیستند: {len(not_found)}")

        if warnings:
            print(f"\n⚠️  {len(warnings)} آیتم نیاز به درج سطر دارند:")
            for w in warnings[:5]:
                print(f"   - {w['item']} ({w['mohor']}): کمبود {w['deficit']} سطر")
            if len(warnings) > 5:
                print(f"   ... و {len(warnings) - 5} مورد دیگر")

        if not_found:
            print(f"\n❌ آیتم‌های قابل درج نیستند:")
            for item in not_found[:5]:
                print(f"   - {item['item']} ({item['mohor']}): {item['reason']}")
            if len(not_found) > 5:
                print(f"   ... و {len(not_found) - 5} مورد دیگر")

        return updates, not_found, warnings

    def _match_item(self, pms_file: str, pms_sheet: str, mohor_name: str,
                    mohor_num: int, pnt_data: Dict, item_locations: Dict,
                    g2_value: Any) -> Dict:
        """
        تطابق یک آیتم PNT با PMS

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت
            mohor_name: نام محور
            mohor_num: شماره محور
            pnt_data: داده‌های آیتم از PNT
            item_locations: دیکشنری موقعیت‌ها
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه تطابق
        """
        normalized_item = pnt_data['normalized']

        # بررسی وجود در PMS
        if normalized_item in item_locations:
            mohor_locations = [
                loc for loc in item_locations[normalized_item]
                if loc['mohor'] == mohor_name
            ]

            if mohor_locations:
                # آیتم موجود است
                return self._create_existing_update(
                    mohor_name, pnt_data, mohor_locations, g2_value
                )

        # آیتم جدید - باید درج شود
        return self._create_new_update(
            pms_file, pms_sheet, mohor_name, mohor_num, pnt_data, g2_value
        )

    def _create_existing_update(self, mohor_name: str, pnt_data: Dict,
                                locations: List[Dict], g2_value: Any) -> Dict:
        """
        ایجاد به‌روزرسانی برای آیتم موجود

        Args:
            mohor_name: نام محور
            pnt_data: داده‌های PNT
            locations: لیست موقعیت‌های آیتم در PMS
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه
        """
        needed_quantity = pnt_data['quantity']
        current_quantity = len(locations)

        result = {
            'status': 'existing',
            'update': {
                'mohor': mohor_name,
                'item_text': pnt_data['single_line'],
                'existing_rows': [loc['row'] for loc in locations],
                'needed_quantity': needed_quantity,
                'a_value': pnt_data['single_line'],
                'e_value': g2_value,
                'n_value': pnt_data['m_value'],
                'is_new_item': False
            }
        }

        # بررسی کمبود
        if current_quantity < needed_quantity:
            deficit = needed_quantity - current_quantity
            result['warning'] = {
                'item': pnt_data['single_line'],
                'mohor': mohor_name,
                'needed': needed_quantity,
                'available': current_quantity,
                'deficit': deficit
            }

        return result

    def _create_new_update(self, pms_file: str, pms_sheet: str,
                           mohor_name: str, mohor_num: int,
                           pnt_data: Dict, g2_value: Any) -> Dict:
        """
        ایجاد به‌روزرسانی برای آیتم جدید

        Args:
            pms_file: مسیر فایل PMS
            pms_sheet: نام شیت
            mohor_name: نام محور
            mohor_num: شماره محور
            pnt_data: داده‌های PNT
            g2_value: مقدار G2

        Returns:
            دیکشنری نتیجه
        """
        last_level5 = self.hierarchy_searcher.find_last_level5_in_section(
            pms_file, pms_sheet, mohor_num
        )

        if last_level5:
            return {
                'status': 'new',
                'update': {
                    'mohor': mohor_name,
                    'item_text': f"🆕 {pnt_data['single_line']}",
                    'existing_rows': [last_level5],
                    'needed_quantity': pnt_data['quantity'],
                    'a_value': pnt_data['single_line'],
                    'e_value': g2_value,
                    'n_value': pnt_data['m_value'],
                    'g_value': pnt_data['m_value'],  # ⬅️ برای ستون G (آیتم جدید)
                    'is_new_item': True
                }
            }
        else:
            return {
                'status': 'not_found',
                'error': {
                    'item': pnt_data['single_line'],
                    'mohor': mohor_name,
                    'reason': 'محور یا Level 5 در PMS یافت نشد'
                }
            }


# ================================================================================
# 🔄 ماژول به‌روزرسانی با COM
# ================================================================================

class COMUpdater:
    """
    کلاس به‌روزرسانی فایل Excel با استفاده از win32com
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config

    def update_file(self, file_path: str, sheet_name: str, updates: List[Dict]):
        """
        به‌روزرسانی فایل PMS

        Args:
            file_path: مسیر فایل
            sheet_name: نام شیت
            updates: لیست به‌روزرسانی‌ها
        """
        print("\n" + "=" * 100)
        print("🔧 باز کردن Excel با COM...")
        print("=" * 100)

        abs_path = os.path.abspath(file_path)
        print(f"📂 مسیر فایل: {abs_path}")

        if not os.path.exists(abs_path):
            raise FileNotFoundError(f"❌ فایل یافت نشد: {abs_path}")

        # باز کردن فایل با openpyxl برای بررسی ستون E
        print("🔍 بررسی وضعیت ستون E...")
        wb_check = openpyxl.load_workbook(abs_path, data_only=True)
        ws_check = wb_check[sheet_name]

        xl = None
        wb = None

        try:
            xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = False
            xl.DisplayAlerts = False

            print(f"🔓 در حال باز کردن Workbook...")
            wb = xl.Workbooks.Open(abs_path)

            if wb is None:
                raise RuntimeError("❌ Workbook باز نشد")

            ws = wb.Worksheets(sheet_name)
            print(f"✅ شیت '{sheet_name}' یافت شد")

            # پردازش به‌روزرسانی‌ها
            stats = self._process_updates(ws, ws_check, updates)

            # ذخیره
            print(f"\n💾 ذخیره‌سازی...")
            wb.Save()

            # گزارش نهایی
            self._print_stats(stats)

        except Exception as e:
            print(f"\n❌ خطا: {e}")
            import traceback
            traceback.print_exc()
            raise

        finally:
            try:
                wb_check.close()
                if wb:
                    wb.Close(SaveChanges=False)
                if xl:
                    xl.Quit()
            except:
                pass

    def _process_updates(self, ws_com, ws_check, updates: List[Dict]) -> Dict:
        """
        پردازش لیست به‌روزرسانی‌ها

        Args:
            ws_com: worksheet COM
            ws_check: worksheet openpyxl (برای چک کردن ستون E)
            updates: لیست به‌روزرسانی‌ها

        Returns:
            دیکشنری آمار
        """
        stats = {
            'inserted': 0,
            'updated': 0,
            'skipped': 0
        }

        for update in updates:
            if update.get('is_new_item'):
                stats['inserted'] += self._process_new_item(ws_com, update)
            else:
                result = self._process_existing_item(ws_com, ws_check, update)
                stats['inserted'] += result['inserted']
                stats['updated'] += result['updated']
                stats['skipped'] += result['skipped']

        return stats

    def _process_new_item(self, ws, update: Dict) -> int:
        """
        پردازش آیتم جدید (درج + کپی کامل از ردیف الگو + آپدیت A, E, G, N)

        Args:
            ws: worksheet COM
            update: دیکشنری به‌روزرسانی

        Returns:
            تعداد ردیف‌های درج شده
        """
        item_text = update['item_text']
        existing_rows = update['existing_rows']
        needed_quantity = update['needed_quantity']
        a_value = update['a_value']
        e_value = update['e_value']
        g_value = update['g_value']
        n_value = update['n_value']

        print(f"\n📝 {item_text}")
        print(f"   🆕 آیتم جدید - درج {needed_quantity} ردیف")

        # ردیف الگو = آخرین ردیف Level 5 پیدا شده
        template_row = existing_rows[-1]
        inserted_count = 0

        for i in range(needed_quantity):
            new_row = template_row + i + 1
            ws.Rows(new_row).Insert()

            # ✅ مرحله 1: کپی کامل ردیف (شامل تمام مقادیر و فرمول‌ها)
            self._copy_row_with_values(ws, template_row, new_row)

            # ✅ مرحله 2: بازنویسی فقط ستون‌های A, E, G, N
            ws.Cells(new_row, self.config.PMS.TEXT_COL).Value = a_value
            ws.Cells(new_row, self.config.PMS.DATE_COL).Value = e_value
            ws.Cells(new_row, self.config.PMS.G_COL).Value = g_value
            ws.Cells(new_row, self.config.PMS.N_COL).Value = n_value

            existing_rows.append(new_row)
            inserted_count += 1

        print(f"   ✅ {needed_quantity} ردیف درج شد")
        return inserted_count

    def _process_existing_item(self, ws_com, ws_check, update: Dict) -> Dict:
        """
        پردازش آیتم موجود (فیلترینگ E + درج در صورت کمبود + آپدیت A, E, N)

        Args:
            ws_com: worksheet COM
            ws_check: worksheet openpyxl
            update: دیکشنری به‌روزرسانی

        Returns:
            دیکشنری آمار {'inserted': ..., 'updated': ..., 'skipped': ...}
        """
        item_text = update['item_text']
        existing_rows = update['existing_rows']
        needed_quantity = update['needed_quantity']
        a_value = update['a_value']
        e_value = update['e_value']
        n_value = update['n_value']

        print(f"\n📝 {item_text}")

        # فیلتر ردیف‌های خالی و پُر
        filled_rows = []
        empty_rows = []

        for row in existing_rows:
            e_cell_value = ws_check.cell(row, self.config.PMS.DATE_COL).value
            if e_cell_value is None or str(e_cell_value).strip() == "":
                empty_rows.append(row)
            else:
                filled_rows.append(row)

        stats = {'inserted': 0, 'updated': 0, 'skipped': len(filled_rows)}

        if filled_rows:
            print(f"   ⏭️  {len(filled_rows)} ردیف پُر (نادیده گرفته شد)")

        print(f"   📊 موجود: {len(existing_rows)} | خالی: {len(empty_rows)} | نیاز: {needed_quantity}")

        # محاسبه کمبود
        deficit = max(0, needed_quantity - len(empty_rows))

        if deficit > 0:
            print(f"   ➕ درج {deficit} ردیف جدید...")
            last_row = existing_rows[-1]

            for i in range(deficit):
                new_row = last_row + i + 1
                ws_com.Rows(new_row).Insert()

                # ✅ کپی کامل ردیف (با مقادیر)
                self._copy_row_with_values(ws_com, last_row, new_row)

                # آپدیت فقط A, E, N (G دست نخورده)
                ws_com.Cells(new_row, self.config.PMS.TEXT_COL).Value = a_value
                ws_com.Cells(new_row, self.config.PMS.DATE_COL).Value = e_value
                ws_com.Cells(new_row, self.config.PMS.N_COL).Value = n_value

                empty_rows.append(new_row)
                existing_rows.append(new_row)
                stats['inserted'] += 1

        # آپدیت ردیف‌های خالی
        rows_to_update = empty_rows[:needed_quantity]

        for row in rows_to_update:
            ws_com.Cells(row, self.config.PMS.TEXT_COL).Value = a_value
            ws_com.Cells(row, self.config.PMS.DATE_COL).Value = e_value
            ws_com.Cells(row, self.config.PMS.N_COL).Value = n_value
            stats['updated'] += 1

        print(f"   ✅ {len(rows_to_update)} ردیف آپدیت شد")
        print(f"   📍 ردیف‌ها: {', '.join(map(str, rows_to_update))}")

        return stats

    def _copy_row_complete(self, ws, source_row: int, target_row: int):
        """
        کپی کامل یک ردیف (محتوا + استایل)

        Args:
            ws: worksheet COM
            source_row: شماره ردیف مبدا
            target_row: شماره ردیف مقصد
        """
        try:
            source_range = ws.Rows(source_row)
            target_range = ws.Rows(target_row)

            source_range.Copy()
            target_range.PasteSpecial(Paste=-4122)  # xlPasteAll

            # تنظیم outline level
            target_range.OutlineLevel = source_range.OutlineLevel

        except Exception as e:
            print(f"⚠️  خطا در کپی ردیف {source_row}: {e}")
            raise

    def _copy_row_with_values(self, ws, source_row: int, target_row: int):
        """
        کپی کامل یک ردیف (محتوا + فرمول + استایل + outline)

        Args:
            ws: worksheet COM
            source_row: شماره ردیف مبدا (الگو)
            target_row: شماره ردیف مقصد (جدید)
        """
        try:
            source_range = ws.Rows(source_row)
            target_range = ws.Rows(target_row)

            # کپی کامل (All = محتوا + فرمت + فرمول)
            source_range.Copy()
            target_range.PasteSpecial(Paste=-4104)  # xlPasteAll

            # تنظیم outline level
            target_range.OutlineLevel = source_range.OutlineLevel

        except Exception as e:
            print(f"⚠️  خطا در کپی ردیف {source_row}: {e}")
            raise

    @staticmethod
    def _print_stats(stats: Dict):
        """
        چاپ آمار نهایی

        Args:
            stats: دیکشنری آمار
        """
        print(f"\n" + "=" * 100)
        print(f"✅ خلاصه:")
        print(f"   🆕 ردیف‌های درج شده: {stats['inserted']}")
        print(f"   🔄 ردیف‌های آپدیت شده: {stats['updated']}")
        print(f"   ⏭️  ردیف‌های نادیده گرفته شده (E پُر): {stats['skipped']}")
        print("=" * 100)


# ================================================================================
# 🚀 هماهنگ‌کننده اصلی
# ================================================================================

class PMSUpdateOrchestrator:
    """
    کلاس هماهنگی کل فرآیند به‌روزرسانی PMS از PNT-G
    """

    def __init__(self, config: PMSConfig = PMSConfig()):
        """
        مقداردهی اولیه

        Args:
            config: تنظیمات برنامه
        """
        self.config = config
        self.cache_manager = PMSCacheManager(config.CACHE_FILE)
        self.structure_reader = PMSStructureReader(config)
        self.pnt_extractor = PNTItemExtractor(config)
        self.update_planner = UpdatePlanner(config)
        self.com_updater = COMUpdater(config)

    def run(self):
        """
        اجرای کامل فرآیند به‌روزرسانی

        Returns:
            دیکشنری نتایج
        """
        print("=" * 100)
        print("🚀 شروع فرآیند کامل به‌روزرسانی PMS از PNT-G")
        print("=" * 100)

        # ══════════════════════════════════════════════════════════════════════
        # مرحله 1: بارگذاری ساختار PMS (با Cache)
        # ══════════════════════════════════════════════════════════════════════
        print("\n📥 بارگذاری ساختار PMS...")
        item_locations = self._load_pms_structure()
        print(f"✅ {len(item_locations)} آیتم منحصر به فرد")

        # ══════════════════════════════════════════════════════════════════════
        # مرحله 2: استخراج آیتم‌های PNT-G
        # ══════════════════════════════════════════════════════════════════════
        print(f"\n🔍 استخراج آیتم‌های PNT-G...")
        items_by_axis, unidentified, g2_value = self.pnt_extractor.extract_all_items(
            self.config.PNT_FILE,
            self.config.PNT_SHEET
        )

        # ══════════════════════════════════════════════════════════════════════
        # مرحله 3: تطابق و برنامه‌ریزی به‌روزرسانی‌ها
        # ══════════════════════════════════════════════════════════════════════
        updates, not_found, warnings = self.update_planner.plan_updates(
            self.config.PMS_FILE,
            self.config.PMS_SHEET,
            item_locations,
            items_by_axis,
            g2_value
        )

        # ══════════════════════════════════════════════════════════════════════
        # مرحله 4: اجرای به‌روزرسانی
        # ══════════════════════════════════════════════════════════════════════
        if updates:
            self.com_updater.update_file(
                self.config.PMS_FILE,
                self.config.PMS_SHEET,
                updates
            )
        else:
            print("\n⚠️  هیچ به‌روزرسانی‌ای برای اجرا وجود ندارد!")

        # ══════════════════════════════════════════════════════════════════════
        # گزارش نهایی
        # ══════════════════════════════════════════════════════════════════════
        self._print_final_report(updates, not_found, warnings, unidentified)

        return {
            'processed': len(updates),
            'not_found': len(not_found),
            'warnings': len(warnings),
            'unidentified_axis': len(unidentified)
        }

    def _load_pms_structure(self) -> Dict:
        """
        بارگذاری ساختار PMS با استفاده از Cache

        Returns:
            دیکشنری موقعیت آیتم‌ها
        """
        # تلاش برای بارگذاری از Cache
        cached_data = self.cache_manager.load_cache(
            self.config.PMS_FILE,
            self.config.PMS_SHEET
        )

        if cached_data:
            print("✅ استفاده از Cache (فایل تغییر نکرده)")
            return cached_data

        # استخراج از فایل
        item_locations = self.structure_reader.extract_all_items(
            self.config.PMS_FILE,
            self.config.PMS_SHEET
        )

        # ذخیره Cache
        self.cache_manager.save_cache(
            self.config.PMS_FILE,
            self.config.PMS_SHEET,
            item_locations
        )

        return item_locations

    @staticmethod
    def _print_final_report(updates, not_found, warnings, unidentified):
        """
        چاپ گزارش نهایی

        Args:
            updates: لیست به‌روزرسانی‌ها
            not_found: لیست آیتم‌های پیدا نشده
            warnings: لیست هشدارها
            unidentified: لیست آیتم‌های بدون محور
        """
        print("\n" + "=" * 100)
        print("🏁 پایان عملیات")
        print("=" * 100)
        print(f"📊 خلاصه گزارش:")

        existing_count = sum(1 for u in updates if not u.get('is_new_item'))
        new_count = sum(1 for u in updates if u.get('is_new_item'))

        print(f"   ✅ آیتم‌های موجود پردازش شده: {existing_count}")
        print(f"   🆕 آیتم‌های جدید درج شده: {new_count}")
        print(f"   ❌ آیتم‌های ناموفق: {len(not_found)}")
        print(f"   ⚠️  آیتم‌های دارای کمبود: {len(warnings)}")

        if unidentified:
            print(f"   🔍 آیتم‌های بدون محور: {len(unidentified)}")

        print("=" * 100)


# ================================================================================
# 🎯 نقطه ورود برنامه
# ================================================================================

def main():
    """
    تابع اصلی برنامه
    """
    try:
        # ایجاد orchestrator با تنظیمات پیش‌فرض
        orchestrator = PMSUpdateOrchestrator()

        # اجرای فرآیند
        results = orchestrator.run()

        print("\n✅ عملیات با موفقیت انجام شد!")

        return results

    except FileNotFoundError as e:
        print(f"\n❌ خطای فایل: {e}")
        return None
    except Exception as e:
        print(f"\n❌ خطای کلی: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    main()
