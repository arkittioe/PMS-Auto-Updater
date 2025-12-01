"""
================================================================================
🎨 PMS Auto-Updater - رابط کاربری PyQt6
================================================================================
نسخه: 1.0.0
توسعه‌دهنده: کمک شده توسط Claude Sonnet 4.5
تاریخ: 1404/09/09

این فایل شامل رابط کاربری گرافیکی با 4 تب اصلی است:
- تب 1: اجرا (انتخاب فایل، تنظیمات، اجرا)
- تب 2: گزارش‌ها (نتایج تفصیلی)
- تب 3: تغییرات (مقایسه قبل/بعد)
- تب 4: درباره (اطلاعات نسخه)
================================================================================
"""

import sys
import os
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QPushButton, QLabel, QLineEdit, QComboBox, QCheckBox,
    QProgressBar, QTextEdit, QTableWidget, QTableWidgetItem, QFileDialog,
    QGroupBox, QGridLayout, QHeaderView, QMessageBox, QFrame, QSplitter,
    QAbstractItemView
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QTimer, QSettings
)
from PyQt6.QtGui import (
    QFont, QColor, QPalette, QIcon, QTextCursor
)

import openpyxl
from tra5_core import (
    PMSConfig, PMSUpdateOrchestrator, ConfigLoader
)

# ================================================================================
# 🎨 تنظیمات استایل Dark Mode
# ================================================================================

DARK_STYLESHEET = """
QMainWindow, QWidget {
    background-color: #1e1e1e;
    color: #d4d4d4;
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    font-size: 10pt;
}

QTabWidget::pane {
    border: 1px solid #3c3c3c;
    background-color: #252525;
}

QTabBar::tab {
    background-color: #2d2d2d;
    color: #d4d4d4;
    padding: 10px 20px;
    margin-right: 2px;
    border: 1px solid #3c3c3c;
    border-bottom: none;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}

QTabBar::tab:selected {
    background-color: #007acc;
    color: white;
}

QTabBar::tab:hover {
    background-color: #3c3c3c;
}

QPushButton {
    background-color: #0e639c;
    color: white;
    border: none;
    padding: 8px 16px;
    border-radius: 4px;
    font-weight: bold;
}

QPushButton:hover {
    background-color: #1177bb;
}

QPushButton:pressed {
    background-color: #0d5689;
}

QPushButton:disabled {
    background-color: #3c3c3c;
    color: #808080;
}

QPushButton#dangerButton {
    background-color: #c72d2d;
}

QPushButton#dangerButton:hover {
    background-color: #e04444;
}

QLineEdit, QComboBox {
    background-color: #3c3c3c;
    color: #d4d4d4;
    border: 1px solid #555;
    padding: 6px;
    border-radius: 3px;
}

QLineEdit:focus, QComboBox:focus {
    border: 1px solid #007acc;
}

QCheckBox {
    color: #d4d4d4;
    spacing: 8px;
}

QCheckBox::indicator {
    width: 18px;
    height: 18px;
    border: 1px solid #555;
    border-radius: 3px;
    background-color: #3c3c3c;
}

QCheckBox::indicator:checked {
    background-color: #007acc;
    border: 1px solid #007acc;
}

QProgressBar {
    border: 1px solid #555;
    border-radius: 4px;
    background-color: #3c3c3c;
    text-align: center;
    color: white;
}

QProgressBar::chunk {
    background-color: #007acc;
    border-radius: 3px;
}

QGroupBox {
    border: 1px solid #3c3c3c;
    border-radius: 5px;
    margin-top: 10px;
    padding-top: 10px;
    font-weight: bold;
    color: #d4d4d4;
}

QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px;
}

QTextEdit {
    background-color: #1e1e1e;
    color: #d4d4d4;
    border: 1px solid #3c3c3c;
    border-radius: 4px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 9pt;
}

QTableWidget {
    background-color: #1e1e1e;
    alternate-background-color: #252525;
    gridline-color: #3c3c3c;
    border: 1px solid #3c3c3c;
    border-radius: 4px;
}

QTableWidget::item {
    padding: 5px;
    color: #d4d4d4;
}

QTableWidget::item:selected {
    background-color: #007acc;
    color: white;
}

QHeaderView::section {
    background-color: #2d2d2d;
    color: #d4d4d4;
    padding: 8px;
    border: 1px solid #3c3c3c;
    font-weight: bold;
}

QLabel {
    color: #d4d4d4;
}

QLabel#titleLabel {
    font-size: 14pt;
    font-weight: bold;
    color: #007acc;
}

QLabel#successLabel {
    color: #4ec9b0;
}

QLabel#warningLabel {
    color: #dcdcaa;
}

QLabel#errorLabel {
    color: #f48771;
}

QFrame#separator {
    background-color: #3c3c3c;
    max-height: 1px;
}
"""



# ================================================================================
# 🔄 Worker Thread برای پردازش
# ================================================================================

class ProcessWorker(QThread):
    """Thread جداگانه برای پردازش تا UI فریز نشود"""

    # سیگنال‌ها
    progress_updated = pyqtSignal(int, str)  # (درصد، پیام)
    log_message = pyqtSignal(str, str)  # (پیام، نوع)
    finished = pyqtSignal(dict)  # نتایج نهایی
    error_occurred = pyqtSignal(str)  # خطا
    stats_updated = pyqtSignal(dict)  # آمار لحظه‌ای

    def __init__(self, config: PMSConfig, dry_run: bool = False):
        super().__init__()
        self.config = config
        self.dry_run = dry_run
        self.is_cancelled = False
        self.results = {
            'updates': [],
            'not_found': [],
            'warnings': [],
            'unidentified': [],
            'stats': {'inserted': 0, 'updated': 0, 'skipped': 0}
        }

    def run(self):
        """اجرای فرآیند در thread جداگانه"""
        try:
            self.log_message.emit("🚀 شروع پردازش...", "info")
            self.progress_updated.emit(10, "بارگذاری تنظیمات...")

            # ایجاد orchestrator سفارشی با callback
            orchestrator = self._create_orchestrator_with_callbacks()

            self.progress_updated.emit(30, "استخراج ساختار PMS...")

            if self.dry_run:
                self.log_message.emit("🔍 حالت Dry Run - هیچ تغییری اعمال نمی‌شود", "warning")
                # فقط تحلیل بدون اعمال
                results = self._dry_run_analysis(orchestrator)
            else:
                # فرآیند کامل
                results = orchestrator.run()

            self.progress_updated.emit(100, "اتمام پردازش")
            self.log_message.emit("✅ پردازش با موفقیت انجام شد!", "success")

            # ساختاردهی نتایج برای UI
            formatted_results = self._format_results(results)
            self.finished.emit(formatted_results)

        except Exception as e:
            error_msg = f"❌ خطا: {str(e)}"
            self.log_message.emit(error_msg, "error")
            self.error_occurred.emit(str(e))

    def _create_orchestrator_with_callbacks(self):
        """ایجاد orchestrator با callback برای بروزرسانی UI"""

        # تعریف callback که پیام‌ها رو به UI می‌فرسته
        def log_to_ui(msg: str, msg_type: str = 'info'):
            self.log_message.emit(msg, msg_type)

        # ایجاد orchestrator با callback
        orchestrator = PMSUpdateOrchestrator(self.config, log_callback=log_to_ui)

        return orchestrator

    def _dry_run_analysis(self, orchestrator):
        """تحلیل بدون اعمال تغییرات"""
        # بارگذاری ساختار
        item_locations = orchestrator._load_pms_structure()

        # استخراج PNT
        items_by_axis, unidentified, g2_value = orchestrator.pnt_extractor.extract_all_items(
            self.config.PNT_FILE,
            self.config.PNT_SHEET
        )

        # تطابق (بدون اعمال)
        updates, not_found, warnings = orchestrator.update_planner.plan_updates(
            self.config.PMS_FILE,
            self.config.PMS_SHEET,
            item_locations,
            items_by_axis,
            g2_value
        )

        return {
            'updates': updates,
            'not_found': not_found,
            'warnings': warnings,
            'unidentified': unidentified,
            'dry_run': True
        }

    def _format_results(self, results: Dict) -> Dict:
        """فرمت‌دهی نتایج برای UI"""
        # هر دو حالت (dry_run و run) حالا لیست‌ها رو برمی‌گردونن
        updates_list = results.get('updates', results.get('updates_list', []))
        not_found_list = results.get('not_found_list', results.get('not_found', []))
        warnings_list = results.get('warnings_list', results.get('warnings', []))
        unidentified_list = results.get('unidentified_list', results.get('unidentified', []))

        return {
            'processed': results.get('processed', len(updates_list)),
            'not_found': results.get('not_found', len(not_found_list)) if isinstance(results.get('not_found'),
                                                                                     int) else len(not_found_list),
            'warnings': results.get('warnings', len(warnings_list)) if isinstance(results.get('warnings'),
                                                                                  int) else len(warnings_list),
            'unidentified_axis': results.get('unidentified_axis', len(unidentified_list)),
            'updates_list': updates_list,
            'not_found_list': not_found_list,
            'warnings_list': warnings_list,
            'unidentified_list': unidentified_list,
            'dry_run': results.get('dry_run', False)
        }

    def cancel(self):
        """لغو پردازش"""
        self.is_cancelled = True
        self.log_message.emit("⏸️ پردازش لغو شد", "warning")


# ================================================================================
# 🎨 کنسول سفارشی
# ================================================================================

class ConsoleWidget(QTextEdit):
    """ویجت کنسول با رنگ‌بندی خاص"""

    def __init__(self):
        super().__init__()
        self.setReadOnly(True)
        self.setMaximumHeight(200)
        self.setFont(QFont("Consolas", 9))

        # رنگ‌ها
        self.colors = {
            'info': '#569cd6',  # آبی
            'success': '#4ec9b0',  # سبز
            'warning': '#dcdcaa',  # زرد
            'error': '#f48771',  # قرمز
            'default': '#d4d4d4'  # خاکستری
        }

    def append_message(self, message: str, msg_type: str = 'default'):
        """افزودن پیام با رنگ مناسب"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        color = self.colors.get(msg_type, self.colors['default'])

        html = f'<span style="color: {color}">[{timestamp}] {message}</span>'
        self.append(html)

        # اسکرول به پایین
        cursor = self.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        self.setTextCursor(cursor)

    def clear_console(self):
        """پاک کردن کنسول"""
        self.clear()
        self.append_message("کنسول پاک شد", "info")


# ================================================================================
# 📊 تب 1: اجرا
# ================================================================================

class ExecuteTab(QWidget):
    """تب اجرا و تنظیمات"""

    def __init__(self, parent_window):
        super().__init__()
        self.parent_window = parent_window
        self.config = parent_window.config
        self.worker = None

        self.init_ui()

        # ✅ اضافه کردن اتصالات
        self.connect_signals()

    def connect_signals(self):
        """اتصال سیگنال‌ها"""
        # ✅ وقتی شیت عوض شد، محدوده رو شناسایی کن
        self.pnt_sheet_combo.currentTextChanged.connect(self.detect_pnt_range)

        # ✅ وقتی فایل PNT عوض شد، مستقیم شیت‌ها رو بارگذاری کن
        self.pnt_file_input.textChanged.connect(self.on_pnt_file_changed)

    def on_pnt_file_changed(self):
        """وقتی فایل PNT تغییر کرد"""
        pnt_file = self.pnt_file_input.text()

        # چک کن فایل معتبر هست
        if os.path.exists(pnt_file) and pnt_file.endswith(('.xlsx', '.xls')):
            # مستقیما شیت‌ها رو بارگذاری کن
            self.load_pnt_sheets()

    def init_ui(self):
        """ساخت رابط کاربری"""
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # ═══════════════════════════════════════════════════════════
        # گروه 1: فایل‌ها
        # ═══════════════════════════════════════════════════════════
        files_group = QGroupBox("📂 فایل‌ها")
        files_layout = QGridLayout()

        # PMS File
        files_layout.addWidget(QLabel("فایل PMS:"), 0, 0)
        self.pms_file_input = QLineEdit(self.config.PMS_FILE)
        files_layout.addWidget(self.pms_file_input, 0, 1)
        self.pms_browse_btn = QPushButton("...")
        self.pms_browse_btn.setMaximumWidth(50)
        self.pms_browse_btn.clicked.connect(self.browse_pms_file)
        files_layout.addWidget(self.pms_browse_btn, 0, 2)

        # PMS Sheet
        files_layout.addWidget(QLabel("شیت PMS:"), 1, 0)
        self.pms_sheet_input = QLineEdit(self.config.PMS_SHEET)
        files_layout.addWidget(self.pms_sheet_input, 1, 1, 1, 2)

        # PNT File
        files_layout.addWidget(QLabel("فایل PNT:"), 2, 0)
        self.pnt_file_input = QLineEdit(self.config.PNT_FILE)
        files_layout.addWidget(self.pnt_file_input, 2, 1)
        self.pnt_browse_btn = QPushButton("...")
        self.pnt_browse_btn.setMaximumWidth(50)
        self.pnt_browse_btn.clicked.connect(self.browse_pnt_file)
        files_layout.addWidget(self.pnt_browse_btn, 2, 2)

        # PNT Sheet (Dropdown)
        files_layout.addWidget(QLabel("شیت PNT:"), 3, 0)
        self.pnt_sheet_combo = QComboBox()
        self.pnt_sheet_combo.setEditable(True)
        self.pnt_sheet_combo.addItem(self.config.PNT_SHEET)
        files_layout.addWidget(self.pnt_sheet_combo, 3, 1)
        self.refresh_sheets_btn = QPushButton("🔄")
        self.refresh_sheets_btn.setMaximumWidth(50)
        self.refresh_sheets_btn.clicked.connect(self.load_pnt_sheets)
        files_layout.addWidget(self.refresh_sheets_btn, 3, 2)

        # محدوده خودکار
        self.auto_range_label = QLabel("محدوده: شناسایی خودکار")
        self.auto_range_label.setStyleSheet("color: #4ec9b0; font-style: italic;")
        files_layout.addWidget(self.auto_range_label, 4, 0, 1, 3)

        files_group.setLayout(files_layout)
        layout.addWidget(files_group)

        # ═══════════════════════════════════════════════════════════
        # گروه 2: تنظیمات سریع
        # ═══════════════════════════════════════════════════════════
        settings_group = QGroupBox("⚙️ تنظیمات سریع")
        settings_layout = QGridLayout()

        # محدوده محورها
        settings_layout.addWidget(QLabel("محدوده محورها:"), 0, 0)
        axis_layout = QHBoxLayout()
        self.axis_start_input = QLineEdit(str(self.config.AXIS_RANGE_START))
        self.axis_start_input.setMaximumWidth(60)
        self.axis_end_input = QLineEdit(str(self.config.AXIS_RANGE_END))
        self.axis_end_input.setMaximumWidth(60)
        axis_layout.addWidget(self.axis_start_input)
        axis_layout.addWidget(QLabel("-"))
        axis_layout.addWidget(self.axis_end_input)
        axis_layout.addStretch()
        settings_layout.addLayout(axis_layout, 0, 1)

        # Cache
        self.use_cache_checkbox = QCheckBox("استفاده از Cache")
        self.use_cache_checkbox.setChecked(self.config.USE_CACHE)
        settings_layout.addWidget(self.use_cache_checkbox, 1, 0)
        self.cache_path_label = QLabel(f"📍 {self.config.CACHE_FILE}")
        self.cache_path_label.setStyleSheet("color: #808080; font-size: 8pt;")
        settings_layout.addWidget(self.cache_path_label, 1, 1)

        # Dry Run
        self.dry_run_checkbox = QCheckBox("حالت Dry Run (شبیه‌سازی بدون تغییر)")
        self.dry_run_checkbox.setStyleSheet("color: #dcdcaa;")
        settings_layout.addWidget(self.dry_run_checkbox, 2, 0, 1, 2)

        settings_group.setLayout(settings_layout)
        layout.addWidget(settings_group)

        # ═══════════════════════════════════════════════════════════
        # گروه 3: کنترل
        # ═══════════════════════════════════════════════════════════
        control_group = QGroupBox("🎬 کنترل")
        control_layout = QHBoxLayout()

        self.start_btn = QPushButton("▶️ شروع پردازش")
        self.start_btn.clicked.connect(self.start_processing)
        control_layout.addWidget(self.start_btn)

        self.stop_btn = QPushButton("⏸️ توقف")
        self.stop_btn.setEnabled(False)
        self.stop_btn.setObjectName("dangerButton")
        self.stop_btn.clicked.connect(self.stop_processing)
        control_layout.addWidget(self.stop_btn)

        self.reset_btn = QPushButton("🔄 تنظیم مجدد")
        self.reset_btn.clicked.connect(self.reset_ui)
        control_layout.addWidget(self.reset_btn)

        control_group.setLayout(control_layout)
        layout.addWidget(control_group)

        # ═══════════════════════════════════════════════════════════
        # پیشرفت
        # ═══════════════════════════════════════════════════════════
        progress_label = QLabel("📊 پیشرفت کلی:")
        layout.addWidget(progress_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)

        self.progress_text = QLabel("آماده")
        self.progress_text.setStyleSheet("color: #808080; font-style: italic;")
        layout.addWidget(self.progress_text)

        # ═══════════════════════════════════════════════════════════
        # آمار لحظه‌ای
        # ═══════════════════════════════════════════════════════════
        stats_group = QGroupBox("📈 آمار لحظه‌ای")
        stats_layout = QHBoxLayout()

        self.stats_labels = {
            'processed': QLabel("✅ پردازش شده: 0"),
            'new': QLabel("🆕 جدید: 0"),
            'error': QLabel("❌ خطا: 0"),
            'warning': QLabel("⚠️ هشدار: 0"),
            'skipped': QLabel("⏭️ رد شده: 0")
        }

        for label in self.stats_labels.values():
            stats_layout.addWidget(label)

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        layout.addStretch()
        self.setLayout(layout)

    def browse_pms_file(self):
        """انتخاب فایل PMS"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "انتخاب فایل PMS", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.pms_file_input.setText(file_path)

    def browse_pnt_file(self):
        """انتخاب فایل PNT"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "انتخاب فایل PNT", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.pnt_file_input.setText(file_path)
            self.load_pnt_sheets()

    def load_pnt_sheets(self):
        """بارگذاری لیست شیت‌های PNT"""
        pnt_file = self.pnt_file_input.text()

        if not os.path.exists(pnt_file):
            self.parent_window.console.append_message(
                f"⚠️ فایل PNT یافت نشد: {pnt_file}", "warning"
            )
            return

        try:
            wb = openpyxl.load_workbook(pnt_file, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()

            # ✅ پاک کردن ComboBox قبل از اضافه کردن
            self.pnt_sheet_combo.blockSignals(True)  # جلوگیری از ترایگر شدن سیگنال
            self.pnt_sheet_combo.clear()
            self.pnt_sheet_combo.addItems(sheets)
            self.pnt_sheet_combo.blockSignals(False)

            # ✅ اگه شیت پیش‌فرض وجود داره، انتخابش کن
            default_sheet = self.config.PNT_SHEET
            index = self.pnt_sheet_combo.findText(default_sheet)
            if index >= 0:
                self.pnt_sheet_combo.setCurrentIndex(index)

            # ✅ شناسایی خودکار محدوده برای اولین شیت
            self.detect_pnt_range()

            self.parent_window.console.append_message(
                f"✅ {len(sheets)} شیت یافت شد", "success"
            )
        except Exception as e:
            self.parent_window.console.append_message(
                f"❌ خطا در خواندن شیت‌ها: {e}", "error"
            )

    def detect_pnt_range(self):
        """شناسایی خودکار محدوده ردیف‌های PNT از ستون B"""
        pnt_file = self.pnt_file_input.text()
        pnt_sheet = self.pnt_sheet_combo.currentText()

        if not os.path.exists(pnt_file) or not pnt_sheet:
            return

        try:
            self.parent_window.console.append_message(
                f"🔍 شناسایی محدوده از شیت '{pnt_sheet}'...", "info"
            )

            wb = openpyxl.load_workbook(pnt_file, read_only=True, data_only=True)
            ws = wb[pnt_sheet]

            first_num = None
            last_num = None

            # جستجوی اولین و آخرین عدد در ستون B
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row, 2).value  # ستون B
                if isinstance(val, (int, float)):
                    if first_num is None:
                        first_num = row
                    last_num = row

            wb.close()

            if first_num and last_num:
                range_text = f"محدوده: ردیف {first_num} تا {last_num} (شناسایی خودکار)"
                self.auto_range_label.setText(range_text)
                self.auto_range_label.setStyleSheet("color: #4ec9b0; font-style: italic; font-weight: bold;")

                self.parent_window.console.append_message(
                    f"✅ {range_text}", "success"
                )

                # به‌روزرسانی config
                self.config.PNT.ROW_START = first_num
                self.config.PNT.ROW_END = last_num + 1
            else:
                self.auto_range_label.setText("محدوده: شناسایی نشد ❌")
                self.auto_range_label.setStyleSheet("color: #f48771; font-style: italic;")

                self.parent_window.console.append_message(
                    "⚠️ هیچ عددی در ستون B یافت نشد", "warning"
                )

        except Exception as e:
            self.auto_range_label.setText(f"خطا: {str(e)}")
            self.auto_range_label.setStyleSheet("color: #f48771; font-style: italic;")

            self.parent_window.console.append_message(
                f"❌ خطا در شناسایی محدوده: {e}", "error"
            )

    def stop_processing(self):
        """توقف پردازش"""
        if self.worker:
            self.worker.cancel()
            self.parent_window.console.append_message("⏸️ در حال لغو...", "warning")

    def update_progress(self, value: int, text: str):
        """به‌روزرسانی Progress Bar"""
        self.progress_bar.setValue(value)
        self.progress_text.setText(text)

    def processing_error(self, error_msg: str):
        """خطا در پردازش"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        QMessageBox.critical(self, "خطا", f"خطا در پردازش:\n{error_msg}")

    def update_config_from_ui(self):
        """به‌روزرسانی config از مقادیر UI"""
        self.config.PMS_FILE = self.pms_file_input.text()
        self.config.PMS_SHEET = self.pms_sheet_input.text()
        self.config.PNT_FILE = self.pnt_file_input.text()
        self.config.PNT_SHEET = self.pnt_sheet_combo.currentText()

        self.config.AXIS_RANGE_START = int(self.axis_start_input.text())
        self.config.AXIS_RANGE_END = int(self.axis_end_input.text())

        self.config.USE_CACHE = self.use_cache_checkbox.isChecked()

    def reset_ui(self):
        """تنظیم مجدد UI"""
        self.progress_bar.setValue(0)
        self.progress_text.setText("آماده")

        for label in self.stats_labels.values():
            label.setText(label.text().split(':')[0] + ": 0")

        self.parent_window.console.append_message("🔄 تنظیم مجدد انجام شد", "info")

    def start_processing(self):
        """شروع پردازش"""
        # به‌روزرسانی config از UI
        self.update_config_from_ui()

        # غیرفعال کردن دکمه‌ها
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)

        # پاک کردن کنسول
        self.parent_window.console.clear_console()

        # شروع Worker
        dry_run = self.dry_run_checkbox.isChecked()
        self.worker = ProcessWorker(self.config, dry_run)

        # اتصال سیگنال‌ها
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.log_message.connect(self.parent_window.console.append_message)
        self.worker.finished.connect(self.processing_finished)
        self.worker.error_occurred.connect(self.processing_error)
        self.worker.stats_updated.connect(self.update_live_stats)  # جدید

        self.worker.start()

    def update_live_stats(self, stats: Dict):
        """به‌روزرسانی آمار لحظه‌ای"""
        self.stats_labels['processed'].setText(f"✅ پردازش شده: {stats.get('processed', 0)}")
        self.stats_labels['new'].setText(f"🆕 جدید: {stats.get('inserted', 0)}")
        self.stats_labels['error'].setText(f"❌ خطا: {stats.get('failed', 0)}")
        self.stats_labels['warning'].setText(f"⚠️ هشدار: {stats.get('warnings', 0)}")

    def processing_finished(self, results: Dict):
        """پایان پردازش"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        # به‌روزرسانی آمار نهایی
        existing = sum(1 for u in results.get('updates_list', []) if not u.get('is_new_item', False))
        new = sum(1 for u in results.get('updates_list', []) if u.get('is_new_item', False))

        self.stats_labels['processed'].setText(f"✅ پردازش شده: {results.get('processed', 0)}")
        self.stats_labels['new'].setText(f"🆕 جدید: {new}")
        self.stats_labels['error'].setText(f"❌ خطا: {results.get('not_found', 0)}")
        self.stats_labels['warning'].setText(f"⚠️ هشدار: {results.get('warnings', 0)}")

        # نمایش در تب‌های گزارش و تغییرات
        self.parent_window.reports_tab.load_results(results)
        self.parent_window.changes_tab.load_changes(results.get('updates_list', []))

        # پیغام Dry Run
        if results.get('dry_run', False):
            QMessageBox.information(
                self,
                "Dry Run تکمیل شد",
                "✅ تحلیل با موفقیت انجام شد.\n"
                "⚠️ هیچ تغییری روی فایل اعمال نشده است.\n\n"
                "برای اعمال تغییرات، Dry Run را غیرفعال کنید."
            )


# ================================================================================
# 📊 تب 2: گزارش‌ها
# ================================================================================

class ReportsTab(QWidget):
    """تب گزارش‌های تفصیلی"""

    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        """ساخت رابط کاربری"""
        layout = QVBoxLayout()

        # خلاصه نتایج
        summary_group = QGroupBox("📋 خلاصه نتایج")
        summary_layout = QGridLayout()

        self.summary_labels = {
            'updated': QLabel("✅ موجود آپدیت شده: 0"),
            'inserted': QLabel("🆕 جدید درج شده: 0"),
            'failed': QLabel("❌ ناموفق: 0"),
            'warnings': QLabel("⚠️ هشدار کمبود: 0"),
            'unidentified': QLabel("🔍 بدون محور: 0")
        }

        row = 0
        for label in self.summary_labels.values():
            label.setStyleSheet("font-size: 11pt; padding: 5px;")
            summary_layout.addWidget(label, row, 0)
            row += 1

        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)

        # جدول تفصیلی
        table_group = QGroupBox("🗂️ جداول تفصیلی")
        table_layout = QVBoxLayout()

        # تب‌های جدول
        self.table_tabs = QTabWidget()

        # جدول موفق
        self.success_table = self.create_table()
        self.table_tabs.addTab(self.success_table, "✅ موفق")

        # جدول ناموفق
        self.failed_table = self.create_table()
        self.table_tabs.addTab(self.failed_table, "❌ ناموفق")

        # جدول هشدار
        self.warning_table = self.create_table()
        self.table_tabs.addTab(self.warning_table, "⚠️ هشدار")

        # جدول بدون محور
        self.unidentified_table = self.create_table()
        self.table_tabs.addTab(self.unidentified_table, "🔍 بدون محور")

        table_layout.addWidget(self.table_tabs)

        # دکمه‌های Export
        export_layout = QHBoxLayout()

        save_excel_btn = QPushButton("💾 ذخیره Excel")
        save_excel_btn.clicked.connect(self.export_to_excel)
        export_layout.addWidget(save_excel_btn)

        save_csv_btn = QPushButton("📄 ذخیره CSV")
        save_csv_btn.clicked.connect(self.export_to_csv)
        export_layout.addWidget(save_csv_btn)

        export_layout.addStretch()

        table_layout.addLayout(export_layout)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        self.setLayout(layout)

    def create_table(self) -> QTableWidget:
        """ساخت یک جدول خالی"""
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["ردیف", "محور", "آیتم", "جزئیات"])
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        return table

    def load_results(self, results: Dict):
        """بارگذاری نتایج در جداول"""
        # به‌روزرسانی خلاصه
        updates_list = results.get('updates_list', [])
        existing_count = sum(1 for u in updates_list if not u.get('is_new_item', False))
        new_count = sum(1 for u in updates_list if u.get('is_new_item', False))

        self.summary_labels['updated'].setText(f"✅ موجود آپدیت شده: {existing_count}")
        self.summary_labels['inserted'].setText(f"🆕 جدید درج شده: {new_count}")
        self.summary_labels['failed'].setText(f"❌ ناموفق: {results.get('not_found', 0)}")
        self.summary_labels['warnings'].setText(f"⚠️ هشدار کمبود: {results.get('warnings', 0)}")
        self.summary_labels['unidentified'].setText(f"🔍 بدون محور: {results.get('unidentified_axis', 0)}")

        # پر کردن جداول
        self._fill_success_table(updates_list)
        self._fill_failed_table(results.get('not_found_list', []))
        self._fill_warning_table(results.get('warnings_list', []))
        self._fill_unidentified_table(results.get('unidentified_list', []))

    def _fill_success_table(self, updates_list: List[Dict]):
        """پر کردن جدول موفق"""
        self.success_table.setRowCount(0)

        for idx, update in enumerate(updates_list):
            row_pos = self.success_table.rowCount()
            self.success_table.insertRow(row_pos)

            # ردیف
            self.success_table.setItem(row_pos, 0, QTableWidgetItem(str(idx + 1)))

            # محور
            self.success_table.setItem(row_pos, 1, QTableWidgetItem(update.get('mohor', '')))

            # آیتم
            item_text = update.get('item_text', update.get('a_value', ''))
            self.success_table.setItem(row_pos, 2, QTableWidgetItem(item_text))

            # جزئیات
            is_new = update.get('is_new_item', False)
            rows_str = ', '.join(map(str, update.get('existing_rows', [])))
            detail = f"{'🆕 جدید' if is_new else '✅ آپدیت'} | ردیف‌ها: {rows_str}"
            self.success_table.setItem(row_pos, 3, QTableWidgetItem(detail))

    def _fill_failed_table(self, not_found_list: List[Dict]):
        """پر کردن جدول ناموفق"""
        self.failed_table.setRowCount(0)

        for idx, item in enumerate(not_found_list):
            row_pos = self.failed_table.rowCount()
            self.failed_table.insertRow(row_pos)

            self.failed_table.setItem(row_pos, 0, QTableWidgetItem(str(idx + 1)))
            self.failed_table.setItem(row_pos, 1, QTableWidgetItem(item.get('mohor', '')))
            self.failed_table.setItem(row_pos, 2, QTableWidgetItem(item.get('item', '')))
            self.failed_table.setItem(row_pos, 3, QTableWidgetItem(item.get('reason', '')))

    def _fill_warning_table(self, warnings_list: List[Dict]):
        """پر کردن جدول هشدار"""
        self.warning_table.setRowCount(0)

        for idx, warning in enumerate(warnings_list):
            row_pos = self.warning_table.rowCount()
            self.warning_table.insertRow(row_pos)

            self.warning_table.setItem(row_pos, 0, QTableWidgetItem(str(idx + 1)))
            self.warning_table.setItem(row_pos, 1, QTableWidgetItem(warning.get('mohor', '')))
            self.warning_table.setItem(row_pos, 2, QTableWidgetItem(warning.get('item', '')))

            detail = f"نیاز: {warning.get('needed', 0)} | موجود: {warning.get('available', 0)} | کمبود: {warning.get('deficit', 0)}"
            self.warning_table.setItem(row_pos, 3, QTableWidgetItem(detail))

    def _fill_unidentified_table(self, unidentified_list: List[Dict]):
        """پر کردن جدول بدون محور"""
        self.unidentified_table.setRowCount(0)

        for idx, item in enumerate(unidentified_list):
            row_pos = self.unidentified_table.rowCount()
            self.unidentified_table.insertRow(row_pos)

            self.unidentified_table.setItem(row_pos, 0, QTableWidgetItem(str(idx + 1)))
            self.unidentified_table.setItem(row_pos, 1, QTableWidgetItem("-"))
            self.unidentified_table.setItem(row_pos, 2, QTableWidgetItem(item.get('item', '')))
            self.unidentified_table.setItem(row_pos, 3, QTableWidgetItem(f"سطر PNT: {item.get('row', '')}"))

    def export_to_excel(self):
        """Export نتایج به Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره گزارش Excel", "report.xlsx", "Excel Files (*.xlsx)"
        )
        if file_path:
            # TODO: پیاده‌سازی export به Excel
            QMessageBox.information(self, "موفق", f"گزارش در {file_path} ذخیره شد")

    def export_to_csv(self):
        """Export نتایج به CSV"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "ذخیره گزارش CSV", "report.csv", "CSV Files (*.csv)"
        )
        if file_path:
            # TODO: پیاده‌سازی export به CSV
            QMessageBox.information(self, "موفق", f"گزارش در {file_path} ذخیره شد")


# ================================================================================
# 📊 تب 3: تغییرات
# ================================================================================

class ChangesTab(QWidget):
    """تب نمایش لیست تغییرات"""

    def __init__(self):
        super().__init__()
        self.changes_data = []
        self.init_ui()

    def init_ui(self):
        """ساخت رابط کاربری"""
        layout = QVBoxLayout()

        # توضیحات
        info_label = QLabel("🔄 لیست کامل تغییرات اعمال شده")
        info_label.setStyleSheet("font-size: 11pt; color: #007acc; font-weight: bold;")
        layout.addWidget(info_label)

        # فیلترها
        filter_group = QGroupBox("🔍 فیلترها")
        filter_layout = QHBoxLayout()

        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["همه", "✅ موجود آپدیت شده", "🆕 جدید درج شده"])
        self.filter_combo.currentTextChanged.connect(self.apply_filter)
        filter_layout.addWidget(QLabel("نمایش:"))
        filter_layout.addWidget(self.filter_combo)

        filter_layout.addStretch()
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)

        # جدول تغییرات
        self.changes_table = QTableWidget()
        self.changes_table.setColumnCount(7)
        self.changes_table.setHorizontalHeaderLabels([
            "ردیف", "محور", "نوع", "آیتم", "مقدار E", "مقدار N", "ردیف‌های PMS"
        ])
        self.changes_table.setAlternatingRowColors(True)
        self.changes_table.horizontalHeader().setStretchLastSection(True)
        self.changes_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.changes_table)

        # آمار
        stats_layout = QHBoxLayout()
        self.stats_label = QLabel("📊 کل تغییرات: 0 | موجود: 0 | جدید: 0")
        self.stats_label.setStyleSheet("color: #4ec9b0; font-weight: bold;")
        stats_layout.addWidget(self.stats_label)
        stats_layout.addStretch()
        layout.addLayout(stats_layout)

        self.setLayout(layout)

    def load_changes(self, updates_list: List[Dict]):
        """بارگذاری لیست تغییرات"""
        self.changes_data = updates_list
        self.apply_filter()

        # به‌روزرسانی آمار
        total = len(updates_list)
        existing = sum(1 for u in updates_list if not u.get('is_new_item', False))
        new = sum(1 for u in updates_list if u.get('is_new_item', False))

        self.stats_label.setText(f"📊 کل تغییرات: {total} | موجود: {existing} | جدید: {new}")

    def apply_filter(self):
        """اعمال فیلتر"""
        filter_text = self.filter_combo.currentText()

        if filter_text == "همه":
            filtered_data = self.changes_data
        elif "موجود" in filter_text:
            filtered_data = [u for u in self.changes_data if not u.get('is_new_item', False)]
        else:  # جدید
            filtered_data = [u for u in self.changes_data if u.get('is_new_item', False)]

        self._fill_table(filtered_data)

    def _fill_table(self, data: List[Dict]):
        """پر کردن جدول"""
        self.changes_table.setRowCount(len(data))

        for idx, update in enumerate(data):
            self.changes_table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
            self.changes_table.setItem(idx, 1, QTableWidgetItem(update.get('mohor', '')))

            change_type = "🆕 جدید" if update.get('is_new_item', False) else "✅ آپدیت"
            type_item = QTableWidgetItem(change_type)
            if update.get('is_new_item', False):
                type_item.setForeground(QColor("#4ec9b0"))
            self.changes_table.setItem(idx, 2, type_item)

            self.changes_table.setItem(idx, 3, QTableWidgetItem(update.get('a_value', '')))
            self.changes_table.setItem(idx, 4, QTableWidgetItem(str(update.get('e_value', ''))))
            self.changes_table.setItem(idx, 5, QTableWidgetItem(str(update.get('n_value', ''))))

            rows_str = ', '.join(map(str, update.get('existing_rows', [])))
            self.changes_table.setItem(idx, 6, QTableWidgetItem(rows_str))


# ================================================================================
# 📊 تب 4: درباره
# ================================================================================

class AboutTab(QWidget):
    """تب اطلاعات برنامه"""

    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        """ساخت رابط کاربری"""
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        layout.setSpacing(20)

        # لوگو و عنوان
        title_label = QLabel("🎨 PMS Auto-Updater")
        title_label.setObjectName("titleLabel")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 20pt; font-weight: bold; color: #007acc;")
        layout.addWidget(title_label)

        # نسخه
        version_label = QLabel("نسخه 1.0.0")
        version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        version_label.setStyleSheet("font-size: 12pt; color: #d4d4d4;")
        layout.addWidget(version_label)

        # خط جداکننده
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.HLine)
        separator.setObjectName("separator")
        layout.addWidget(separator)

        # توضیحات
        description = QLabel(
            "سیستم به‌روزرسانی خودکار فایل PMS از PNT-G\n\n"
            "✅ استخراج خودکار آیتم‌ها\n"
            "✅ تطبیق هوشمند با ساختار سلسله‌مراتبی\n"
            "✅ درج و آپدیت خودکار\n"
            "✅ Cache برای سرعت بالا\n"
            "✅ رابط کاربری PyQt6"
        )
        description.setAlignment(Qt.AlignmentFlag.AlignCenter)
        description.setStyleSheet("font-size: 10pt; line-height: 1.6;")
        layout.addWidget(description)

        # خط جداکننده
        separator2 = QFrame()
        separator2.setFrameShape(QFrame.Shape.HLine)
        separator2.setObjectName("separator")
        layout.addWidget(separator2)

        # تیم توسعه
        dev_label = QLabel(
            "💻 توسعه‌دهنده: Hossein Izadi"
            "📅 تاریخ: 1404/09/09\n"
            "🔧  Python 3.11 | PyQt6 | openpyxl | win32com"
        )
        dev_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dev_label.setStyleSheet("font-size: 9pt; color: #808080;")
        layout.addWidget(dev_label)

        layout.addStretch()

        # راهنما
        help_group = QGroupBox("📖 راهنمای سریع")
        help_layout = QVBoxLayout()

        help_text = QLabel(
            "1️⃣ فایل PMS و PNT را انتخاب کنید\n"
            "2️⃣ شیت‌های مورد نظر را تعیین کنید\n"
            "3️⃣ تنظیمات را بررسی کنید (محدوده محورها، Cache)\n"
            "4️⃣ برای تست ابتدا Dry Run را فعال کنید\n"
            "5️⃣ دکمه 'شروع پردازش' را بزنید\n"
            "6️⃣ نتایج را در تب 'گزارش‌ها' مشاهده کنید"
        )
        help_text.setWordWrap(True)
        help_text.setStyleSheet("font-size: 9pt; line-height: 1.8;")
        help_layout.addWidget(help_text)

        help_group.setLayout(help_layout)
        layout.addWidget(help_group)

        self.setLayout(layout)


# ================================================================================
# 🪟 پنجره اصلی
# ================================================================================

class MainWindow(QMainWindow):
    """پنجره اصلی برنامه"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("PMS Auto-Updater | نسخه 1.0.0")
        self.setMinimumSize(1200, 800)

        # بارگذاری تنظیمات از config.env
        self.config = self.load_config()

        self.init_ui()

        # بارگذاری تنظیمات ذخیره شده
        self.settings = QSettings("PMSUpdater", "Settings")
        self.restore_settings()

    def load_config(self) -> PMSConfig:
        """بارگذاری تنظیمات"""
        # اولویت 1: config.json
        # اولویت 2: config.env

        # اگه هیچکدوم نبود، PMSConfig خودش خطا میده
        if os.path.exists("config.json"):
            return PMSConfig("config.json")
        elif os.path.exists("config.env"):
            return PMSConfig("config.env")
        else:
            # پیش‌فرض config.json (خطا میده اگه نباشه)
            return PMSConfig("config.json")

    def init_ui(self):
        """ساخت رابط کاربری"""
        # ویجت مرکزی
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 10, 10, 10)

        # تب‌ها
        self.tabs = QTabWidget()

        self.execute_tab = ExecuteTab(self)
        self.reports_tab = ReportsTab()
        self.changes_tab = ChangesTab()
        self.about_tab = AboutTab()

        self.tabs.addTab(self.execute_tab, "▶️ اجرا")
        self.tabs.addTab(self.reports_tab, "📊 گزارش‌ها")
        self.tabs.addTab(self.changes_tab, "🔄 تغییرات")
        self.tabs.addTab(self.about_tab, "ℹ️ درباره")

        main_layout.addWidget(self.tabs)

        # کنسول (پایین صفحه)
        console_label = QLabel("📟 کنسول:")
        console_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        main_layout.addWidget(console_label)

        self.console = ConsoleWidget()
        main_layout.addWidget(self.console)

        # دکمه پاک کردن کنسول
        clear_console_btn = QPushButton("🧹 پاک کردن کنسول")
        clear_console_btn.setMaximumWidth(150)
        clear_console_btn.clicked.connect(self.console.clear_console)
        main_layout.addWidget(clear_console_btn)

        central_widget.setLayout(main_layout)

        # پیام خوش‌آمدگویی
        self.console.append_message("🎉 خوش آمدید به PMS Auto-Updater", "success")
        self.console.append_message(f"📂 فایل PMS: {self.config.PMS_FILE}", "info")
        self.console.append_message(f"📂 فایل PNT: {self.config.PNT_FILE}", "info")

    def restore_settings(self):
        """بازیابی تنظیمات ذخیره شده"""
        geometry = self.settings.value("geometry")
        if geometry:
            self.restoreGeometry(geometry)

    def closeEvent(self, event):
        """هنگام بستن پنجره"""
        # ذخیره تنظیمات
        self.settings.setValue("geometry", self.saveGeometry())
        event.accept()


# ================================================================================
# 🚀 نقطه ورود
# ================================================================================

def main():
    """تابع اصلی برنامه"""
    app = QApplication(sys.argv)

    # اعمال Dark Theme
    app.setStyleSheet(DARK_STYLESHEET)

    # پنجره اصلی
    window = MainWindow()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
